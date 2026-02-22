"""엑셀 업로드 API 라우터."""

import io
import os
import tempfile
from datetime import date
from typing import List, Optional

import aiofiles
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pydantic import BaseModel, field_validator
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.account import Account
from app.models.keyword import KeywordPool
from app.models.template import CampaignTemplate
from app.services.excel_parser import ExcelParser, CampaignData as ParsedCampaignData
from app.services.campaign_extension import (
    extract_place_id,
    check_extension_eligible,
    ExtensionInfo,
)


router = APIRouter(prefix="/upload", tags=["upload"])


# Pydantic 모델 정의
class CampaignPreviewItem(BaseModel):
    """미리보기용 캠페인 데이터."""

    row_number: int
    agency_name: str
    user_id: str
    start_date: date
    end_date: date
    daily_limit: int
    keywords: List[str]
    keyword_count: int
    place_name: str
    place_url: str
    campaign_type: str
    is_valid: bool
    errors: List[str]

    # 연장 정보
    extension_eligible: bool = False
    existing_campaign_code: Optional[str] = None
    existing_campaign_id: Optional[int] = None
    existing_total_count: Optional[int] = None

    @classmethod
    def from_parsed(
        cls,
        campaign: ParsedCampaignData,
        extension_info: Optional[ExtensionInfo] = None,
    ) -> "CampaignPreviewItem":
        """ParsedCampaignData에서 변환."""
        item = cls(
            row_number=campaign.row_number,
            agency_name=campaign.agency_name,
            user_id=campaign.user_id,
            start_date=campaign.start_date,
            end_date=campaign.end_date,
            daily_limit=campaign.daily_limit,
            keywords=campaign.keywords,
            keyword_count=len(campaign.keywords),
            place_name=campaign.place_name,
            place_url=campaign.place_url,
            campaign_type=campaign.campaign_type,
            is_valid=campaign.is_valid,
            errors=campaign.errors,
        )

        if extension_info:
            item.extension_eligible = extension_info.is_eligible
            item.existing_campaign_code = extension_info.existing_campaign_code
            item.existing_campaign_id = extension_info.existing_campaign_id
            item.existing_total_count = extension_info.existing_total_count

        return item


class PreviewResponse(BaseModel):
    """업로드 미리보기 응답."""

    success: bool
    total_count: int
    valid_count: int
    invalid_count: int
    campaigns: List[CampaignPreviewItem]
    file_errors: List[str]


class CampaignConfirmItem(BaseModel):
    """확정용 캠페인 데이터."""

    agency_name: str
    user_id: str
    start_date: date
    end_date: date
    daily_limit: int
    keywords: List[str]
    place_name: str
    place_url: str
    campaign_type: str
    action: str = "new"  # "new" 또는 "extend"
    existing_campaign_id: Optional[int] = None  # 연장 시 기존 캠페인 DB ID

    @field_validator("campaign_type")
    @classmethod
    def validate_campaign_type(cls, v):
        if not v or not v.strip():
            raise ValueError("캠페인 이름이 비어있습니다")
        return v.strip()

    @field_validator("action")
    @classmethod
    def validate_action(cls, v):
        if v not in ["new", "extend"]:
            raise ValueError("action은 'new' 또는 'extend'여야 합니다")
        return v


class ConfirmRequest(BaseModel):
    """업로드 확정 요청."""

    campaigns: List[CampaignConfirmItem]


class ConfirmResponse(BaseModel):
    """업로드 확정 응답."""

    success: bool
    message: str
    created_count: int
    new_count: int = 0
    extend_count: int = 0
    skipped: List[str] = []
    campaign_ids: List[int] = []


@router.post("/preview", response_model=PreviewResponse)
async def preview_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    엑셀 파일 업로드 후 미리보기 데이터 반환.

    - 엑셀 파일을 파싱하여 캠페인 데이터 추출
    - 각 행의 유효성 검증 수행
    - 각 행마다 연장 가능 여부 체크 (place_id + 총 타수 조건)
    - 에러가 있는 행도 포함하여 전체 결과 반환
    """
    # 파일 확장자 확인
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다"
        )

    # 임시 파일로 저장
    temp_file = None
    try:
        suffix = os.path.splitext(file.filename)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp_file = temp.name
            content = await file.read()
            async with aiofiles.open(temp_file, 'wb') as f:
                await f.write(content)

        # 유효한 사용자 ID 목록 조회
        valid_user_ids = [
            account.user_id
            for account in db.query(Account).filter(Account.is_active == True).all()
        ]

        # 활성 템플릿 이름 목록 조회
        valid_template_names = [
            t.type_name
            for t in db.query(CampaignTemplate).filter(CampaignTemplate.is_active == True).all()
        ]

        # 파싱 수행
        parser = ExcelParser(
            valid_user_ids=valid_user_ids,
            valid_template_names=valid_template_names,
        )
        result = parser.parse(temp_file)

        # 각 행마다 연장 조건 체크 포함하여 응답 생성
        campaigns = []
        for c in result.campaigns:
            extension_info = None

            # 유효한 행만 연장 체크 (에러 있으면 체크 불필요)
            if c.is_valid:
                place_id = extract_place_id(c.place_url)
                if place_id:
                    # total_count 계산: daily_limit * 기간일수
                    days = (c.end_date - c.start_date).days + 1
                    new_total_count = c.daily_limit * days
                    # 같은 계정의 캠페인만 연장 대상
                    acct = db.query(Account).filter(
                        Account.user_id == c.user_id,
                        Account.is_active == True,
                    ).first()
                    extension_info = check_extension_eligible(
                        place_id=place_id,
                        new_total_count=new_total_count,
                        db=db,
                        new_start_date=c.start_date,
                        account_id=acct.id if acct else None,
                    )

            campaigns.append(
                CampaignPreviewItem.from_parsed(c, extension_info)
            )

        return PreviewResponse(
            success=result.success,
            total_count=len(result.campaigns),
            valid_count=len(result.valid_campaigns),
            invalid_count=len(result.invalid_campaigns),
            campaigns=campaigns,
            file_errors=result.errors
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"파일 처리 중 오류가 발생했습니다: {str(e)}"
        )
    finally:
        # 임시 파일 삭제
        if temp_file and os.path.exists(temp_file):
            os.unlink(temp_file)


@router.post("/confirm", response_model=ConfirmResponse)
async def confirm_upload(
    request: ConfirmRequest,
    db: Session = Depends(get_db)
):
    """
    미리보기 확인 후 캠페인 등록 큐에 추가.

    각 캠페인의 action 필드에 따라 분기 처리:
    - "new": 신규 캠페인으로 등록 대기 (status=pending)
    - "extend": 기존 캠페인 연장 대기 (status=pending_extend)
    """
    from app.models.campaign import Campaign

    if not request.campaigns:
        raise HTTPException(
            status_code=400,
            detail="등록할 캠페인이 없습니다"
        )

    created_count = 0
    new_count = 0
    extend_count = 0
    new_campaign_ids = []
    extend_campaign_ids = []
    skipped = []

    # 한글 레거시 상태 포함 (campaign_extension.py와 동일)
    active_statuses = [
        "active", "집행중",
        "daily_exhausted", "일일소진",
        "campaign_exhausted", "전체소진", "캠페인소진",
    ]

    try:
        for item in request.campaigns:
            # Account 조회
            account = db.query(Account).filter(
                Account.user_id == item.user_id,
                Account.is_active == True
            ).first()

            if not account:
                skipped.append(f"{item.place_name or item.place_url}: 계정 '{item.user_id}' 없음")
                continue

            if item.action == "extend" and item.existing_campaign_id:
                # 연장 대상 캠페인 존재 확인
                target = db.query(Campaign).filter(
                    Campaign.id == item.existing_campaign_id,
                    Campaign.status.in_(active_statuses),
                ).first()

                if not target:
                    skipped.append(f"{item.place_name or item.place_url}: 연장 대상 캠페인 없음")
                    continue

                # total_limit 계산
                days = (item.end_date - item.start_date).days + 1
                total_limit = item.daily_limit * days

                # 연장 대기 캠페인 생성
                campaign = Campaign(
                    account_id=account.id,
                    agency_name=item.agency_name,
                    place_name=item.place_name,
                    place_url=item.place_url,
                    place_id=extract_place_id(item.place_url),
                    campaign_type=item.campaign_type,
                    start_date=item.start_date,
                    end_date=item.end_date,
                    daily_limit=item.daily_limit,
                    total_limit=total_limit,
                    original_keywords=",".join(item.keywords),
                    status="pending_extend",
                    extend_target_id=item.existing_campaign_id,
                    registration_step="queued",
                    registration_message="연장 처리 대기 중...",
                )
                db.add(campaign)
                db.flush()
                extend_campaign_ids.append(campaign.id)
                extend_count += 1
            else:
                # total_limit 계산
                days = (item.end_date - item.start_date).days + 1
                total_limit = item.daily_limit * days

                # 신규 캠페인 생성
                campaign = Campaign(
                    account_id=account.id,
                    agency_name=item.agency_name,
                    place_name=item.place_name,
                    place_url=item.place_url,
                    place_id=extract_place_id(item.place_url),
                    campaign_type=item.campaign_type,
                    start_date=item.start_date,
                    end_date=item.end_date,
                    daily_limit=item.daily_limit,
                    total_limit=total_limit,
                    original_keywords=",".join(item.keywords),
                    status="pending",
                    registration_step="queued",
                    registration_message="등록 대기 중...",
                )
                db.add(campaign)
                db.flush()

                # KeywordPool에 키워드 저장
                seen_kw = set()
                for kw in item.keywords:
                    kw = kw.strip()
                    if kw and kw not in seen_kw:
                        db.add(KeywordPool(
                            campaign_id=campaign.id,
                            keyword=kw,
                            is_used=False,
                        ))
                        seen_kw.add(kw)

                new_campaign_ids.append(campaign.id)
                new_count += 1

            created_count += 1

        db.commit()

        # 자동 등록/연장 트리거
        from app.services.auto_registration import (
            trigger_auto_registration,
            trigger_auto_extension,
        )
        if new_campaign_ids:
            trigger_auto_registration(new_campaign_ids)
        if extend_campaign_ids:
            trigger_auto_extension(extend_campaign_ids)

        # 메시지 생성
        parts = []
        if new_count > 0:
            parts.append(f"신규 {new_count}개")
        if extend_count > 0:
            parts.append(f"연장 {extend_count}개")
        message = f"{', '.join(parts)} 캠페인이 등록 대기열에 추가되었습니다"

        all_campaign_ids = new_campaign_ids + extend_campaign_ids
        return ConfirmResponse(
            success=True,
            message=message,
            created_count=created_count,
            new_count=new_count,
            extend_count=extend_count,
            skipped=skipped,
            campaign_ids=all_campaign_ids,
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"캠페인 저장 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/template")
async def download_template(db: Session = Depends(get_db)):
    """
    엑셀 업로드 양식 다운로드.

    - 필수 컬럼 헤더 + 예시 데이터 1행 포함
    - 타입구분 컬럼에 현재 활성 템플릿 이름 목록을 코멘트로 안내
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "캠페인 등록"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["대행사명", "사용자ID", "시작일", "마감일", "일일 한도", "키워드", "플레이스 URL", "캠페인 이름"]
    col_widths = [14, 16, 14, 14, 12, 50, 40, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # 활성 템플릿 이름 조회
    template_names = [
        t.type_name
        for t in db.query(CampaignTemplate).filter(CampaignTemplate.is_active == True).all()
    ]
    template_hint = ", ".join(template_names) if template_names else "(등록된 템플릿 없음)"

    # 예시 데이터
    example = [
        "대행사A",
        "user1",
        date.today().strftime("%Y-%m-%d"),
        date.today().strftime("%Y-%m-%d"),
        50,
        "키워드1, 키워드2, 키워드3",
        "https://m.place.naver.com/place/12345",
        template_names[0] if template_names else "트래픽",
    ]

    example_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = example_fill
        cell.border = thin_border

    # 타입구분 컬럼에 코멘트 추가 (안내용)
    from openpyxl.comments import Comment
    type_cell = ws.cell(row=1, column=8)
    type_cell.comment = Comment(
        f"사용 가능한 타입: {template_hint}",
        "System",
    )

    # 파일 생성
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=campaign_upload_template.xlsx"
        },
    )
