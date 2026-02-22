"""Excel 파싱 모듈."""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook


@dataclass
class CampaignData:
    """파싱된 캠페인 데이터."""

    agency_name: str
    user_id: str
    start_date: date
    end_date: date
    daily_limit: int
    keywords: List[str]
    place_url: str
    campaign_type: str  # 템플릿 이름 (예: '트래픽', '저장하기' 등)
    place_name: str = ""  # 등록 시 플레이스 URL에서 자동 추출
    row_number: int = 0  # 엑셀 행 번호 (에러 표시용)
    errors: List[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        """유효한 데이터인지 확인."""
        return len(self.errors) == 0


@dataclass
class ParseResult:
    """파싱 결과."""

    success: bool
    campaigns: List[CampaignData]
    errors: List[str]  # 파일 레벨 에러

    @property
    def valid_campaigns(self) -> List[CampaignData]:
        """유효한 캠페인만 반환."""
        return [c for c in self.campaigns if c.is_valid]

    @property
    def invalid_campaigns(self) -> List[CampaignData]:
        """에러가 있는 캠페인만 반환."""
        return [c for c in self.campaigns if not c.is_valid]


# 필수 컬럼 정의
REQUIRED_COLUMNS = [
    "대행사명",
    "사용자ID",
    "시작일",
    "마감일",
    "일일 한도",
    "키워드",
    "플레이스 URL",
    "캠페인 이름",
]

# 최소 키워드 개수
MIN_KEYWORD_COUNT = 1


class ExcelParser:
    """엑셀 파일 파서."""

    def __init__(
        self,
        valid_user_ids: Optional[List[str]] = None,
        valid_template_names: Optional[List[str]] = None,
    ):
        """
        Args:
            valid_user_ids: 유효한 사용자 ID 목록 (accounts 테이블에서 조회)
            valid_template_names: 유효한 템플릿 이름 목록 (templates 테이블에서 조회)
        """
        self.valid_user_ids = valid_user_ids or []
        self.valid_template_names = valid_template_names or []

    def parse(self, file_path: str) -> ParseResult:
        """
        엑셀 파일을 파싱하여 캠페인 데이터 추출.

        Args:
            file_path: 엑셀 파일 경로

        Returns:
            ParseResult: 파싱 결과
        """
        file_errors = []
        campaigns = []

        # 파일 존재 확인
        if not Path(file_path).exists():
            return ParseResult(
                success=False,
                campaigns=[],
                errors=[f"파일을 찾을 수 없습니다: {file_path}"]
            )

        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
        except Exception as e:
            return ParseResult(
                success=False,
                campaigns=[],
                errors=[f"엑셀 파일을 읽을 수 없습니다: {str(e)}"]
            )

        # 헤더 행 확인 (첫 번째 행)
        headers = [cell.value for cell in sheet[1]]
        headers = [h.strip() if h else "" for h in headers]

        # 필수 컬럼 존재 확인
        missing_columns = []
        column_indices = {}

        for col_name in REQUIRED_COLUMNS:
            if col_name in headers:
                column_indices[col_name] = headers.index(col_name)
            else:
                missing_columns.append(col_name)

        if missing_columns:
            return ParseResult(
                success=False,
                campaigns=[],
                errors=[f"필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}"]
            )

        # 데이터 행 파싱 (2행부터)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_data = {col: row[idx] for col, idx in column_indices.items()}
            campaign = self._validate_row(row_data, row_num)
            campaigns.append(campaign)

        workbook.close()

        # 전체 결과 판정
        has_errors = any(not c.is_valid for c in campaigns)
        success = len(campaigns) > 0 and not has_errors and len(file_errors) == 0

        return ParseResult(
            success=success,
            campaigns=campaigns,
            errors=file_errors
        )

    def _validate_row(self, row_data: dict, row_num: int) -> CampaignData:
        """
        행 단위 검증.

        Args:
            row_data: 컬럼명-값 딕셔너리
            row_num: 행 번호

        Returns:
            CampaignData: 검증된 캠페인 데이터
        """
        errors = []

        # 대행사명
        agency_name = self._get_string_value(row_data.get("대행사명"))
        if not agency_name:
            errors.append("대행사명이 비어있습니다")

        # 사용자ID
        user_id = self._get_string_value(row_data.get("사용자ID"))
        if not user_id:
            errors.append("사용자ID가 비어있습니다")
        elif self.valid_user_ids and user_id not in self.valid_user_ids:
            errors.append(f"존재하지 않는 사용자ID입니다: {user_id}")

        # 시작일
        start_date = self._parse_date(row_data.get("시작일"))
        if start_date is None:
            errors.append("시작일 형식이 올바르지 않습니다")

        # 마감일
        end_date = self._parse_date(row_data.get("마감일"))
        if end_date is None:
            errors.append("마감일 형식이 올바르지 않습니다")
        elif start_date and end_date and end_date < start_date:
            errors.append("마감일은 시작일 이후여야 합니다")

        # 일일 한도
        daily_limit = self._parse_int(row_data.get("일일 한도"))
        if daily_limit is None:
            errors.append("일일 한도가 올바르지 않습니다")
        elif daily_limit <= 0:
            errors.append("일일 한도는 0보다 커야 합니다")

        # 키워드
        keywords_str = self._get_string_value(row_data.get("키워드"))
        keywords = self.parse_keywords(keywords_str) if keywords_str else []
        if not keywords:
            errors.append("키워드가 비어있습니다")
        elif len(keywords) < MIN_KEYWORD_COUNT:
            errors.append(f"키워드는 최소 {MIN_KEYWORD_COUNT}개 이상이어야 합니다 (현재: {len(keywords)}개)")

        # 플레이스 URL
        place_url = self._get_string_value(row_data.get("플레이스 URL"))
        if not place_url:
            errors.append("플레이스 URL이 비어있습니다")
        elif not self.validate_url(place_url):
            errors.append("플레이스 URL 형식이 올바르지 않습니다 (m.place.naver.com 포함 필요)")

        # 캠페인 이름 (템플릿 이름과 매칭)
        campaign_type = self._get_string_value(row_data.get("캠페인 이름"))
        if not campaign_type:
            errors.append("캠페인 이름이 비어있습니다")
        elif self.valid_template_names and campaign_type not in self.valid_template_names:
            errors.append(f"캠페인 이름은 {self.valid_template_names} 중 하나여야 합니다 (현재: {campaign_type})")

        return CampaignData(
            agency_name=agency_name or "",
            user_id=user_id or "",
            start_date=start_date or date.today(),
            end_date=end_date or date.today(),
            daily_limit=daily_limit or 0,
            keywords=keywords,
            place_name="",
            place_url=place_url or "",
            campaign_type=campaign_type or "",
            row_number=row_num,
            errors=errors
        )

    def validate_url(self, url: str) -> bool:
        """
        플레이스 URL 검증.

        Args:
            url: 검증할 URL

        Returns:
            bool: 유효 여부
        """
        if not url:
            return False

        # m.place.naver.com 포함 확인
        pattern = r"https?://m\.place\.naver\.com/"
        return bool(re.match(pattern, url))

    def parse_keywords(self, keywords_str: str) -> List[str]:
        """
        키워드 문자열을 리스트로 변환.

        Args:
            keywords_str: 쉼표로 구분된 키워드 문자열

        Returns:
            List[str]: 키워드 리스트
        """
        if not keywords_str:
            return []

        # 쉼표로 분리하고 공백 제거
        keywords = [kw.strip() for kw in keywords_str.split(",")]
        # 빈 문자열 제거
        keywords = [kw for kw in keywords if kw]
        return keywords

    def _get_string_value(self, value) -> Optional[str]:
        """셀 값을 문자열로 변환."""
        if value is None:
            return None
        return str(value).strip()

    def _parse_date(self, value) -> Optional[date]:
        """셀 값을 date로 변환."""
        if value is None:
            return None

        # 이미 datetime/date 객체인 경우
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        # 문자열인 경우 파싱 시도
        str_value = str(value).strip()
        date_formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(str_value, fmt).date()
            except ValueError:
                continue

        return None

    def _parse_int(self, value) -> Optional[int]:
        """셀 값을 정수로 변환."""
        if value is None:
            return None

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            return int(value)

        try:
            return int(str(value).strip())
        except ValueError:
            return None
