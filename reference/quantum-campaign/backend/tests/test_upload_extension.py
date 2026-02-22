"""Upload API - 연장/신규 선택 기능 테스트.

Phase 3 - Task 3.5: 대량 등록 미리보기 & 연장/신규 선택
"""

import io
from datetime import date, datetime, timedelta, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.main import app
from app.models.account import Account
from app.models.campaign import Campaign
from app.services.excel_parser import REQUIRED_COLUMNS


# 테스트용 인메모리 DB 설정
SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"

engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture(scope="function")
def db_session():
    """테스트용 DB 세션."""
    Base.metadata.create_all(bind=engine)
    session = TestingSessionLocal()

    # 테스트용 계정 추가
    test_accounts = [
        Account(user_id="testuser1", agency_name="테스트대행사1", is_active=True),
        Account(user_id="testuser2", agency_name="테스트대행사2", is_active=True),
    ]
    for account in test_accounts:
        session.add(account)
    session.commit()

    yield session

    session.close()
    Base.metadata.drop_all(bind=engine)


@pytest.fixture(scope="function")
def client(db_session):
    """테스트 클라이언트."""
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def create_test_excel(rows_data: list) -> bytes:
    """테스트용 엑셀 파일 생성."""
    workbook = Workbook()
    sheet = workbook.active

    for col, header in enumerate(REQUIRED_COLUMNS, 1):
        sheet.cell(row=1, column=col, value=header)

    for row_num, row_data in enumerate(rows_data, start=2):
        for col, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col, value=value)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def make_valid_row(
    agency="테스트대행사1",
    user_id="testuser1",
    place_url="https://m.place.naver.com/restaurant/12345",
    campaign_type="트래픽",
    daily_limit=100,
):
    """유효한 엑셀 데이터 행 생성 헬퍼."""
    tomorrow = date.today() + timedelta(days=1)
    next_week = date.today() + timedelta(days=7)
    keywords = ",".join([f"키워드{i}" for i in range(15)])

    return [
        agency,
        user_id,
        tomorrow.isoformat(),
        next_week.isoformat(),
        daily_limit,
        keywords,
        place_url,
        campaign_type,
    ]


# ============================================================
# Preview API - 연장 정보 포함 테스트
# ============================================================

class TestPreviewExtensionInfo:
    """POST /upload/preview - 연장 정보 포함 테스트."""

    def test_preview_no_existing_campaign(self, client, db_session):
        """기존 캠페인 없으면 extension_eligible = False."""
        rows = [make_valid_row()]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True

        campaign = data["campaigns"][0]
        assert campaign["extension_eligible"] is False
        assert campaign["existing_campaign_code"] is None
        assert campaign["existing_campaign_id"] is None
        assert campaign["existing_total_count"] is None

    def test_preview_with_existing_active_campaign(self, client, db_session):
        """기존 active 캠페인이 있으면 연장 정보 포함."""
        # 기존 active 캠페인 생성
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        existing = Campaign(
            campaign_code="C12345",
            account_id=account.id,
            place_name="테스트플레이스",
            place_url="https://m.place.naver.com/restaurant/12345",
            place_id="12345",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=7),
            end_date=date.today() + timedelta(days=7),
            daily_limit=100,
            total_limit=1400,
            status="active",
        )
        db_session.add(existing)
        db_session.commit()

        # 같은 place_id로 엑셀 업로드
        rows = [make_valid_row(
            place_url="https://m.place.naver.com/restaurant/12345",
            daily_limit=100,
        )]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        campaign = data["campaigns"][0]

        assert campaign["extension_eligible"] is True
        assert campaign["existing_campaign_code"] == "C12345"
        assert campaign["existing_campaign_id"] == existing.id
        assert campaign["existing_total_count"] == 1400

    def test_preview_extension_total_exceeds_limit(self, client, db_session):
        """총 타수 초과 시 extension_eligible = False."""
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        existing = Campaign(
            campaign_code="C99999",
            account_id=account.id,
            place_name="테스트플레이스",
            place_url="https://m.place.naver.com/restaurant/99999",
            place_id="99999",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=7),
            end_date=date.today() + timedelta(days=7),
            daily_limit=500,
            total_limit=9500,  # 9500 + 새 700 = 10200 > 10000
            status="active",
        )
        db_session.add(existing)
        db_session.commit()

        rows = [make_valid_row(
            place_url="https://m.place.naver.com/restaurant/99999",
            daily_limit=100,  # 7일 * 100 = 700
        )]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        campaign = data["campaigns"][0]

        assert campaign["extension_eligible"] is False
        assert campaign["existing_campaign_code"] == "C99999"
        assert campaign["existing_total_count"] == 9500

    def test_preview_extension_pending_campaign_not_eligible(self, client, db_session):
        """pending 상태 캠페인은 연장 대상 아님."""
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        pending = Campaign(
            campaign_code="CPEND",
            account_id=account.id,
            place_name="테스트플레이스",
            place_url="https://m.place.naver.com/restaurant/77777",
            place_id="77777",
            campaign_type="트래픽",
            start_date=date.today(),
            end_date=date.today() + timedelta(days=7),
            daily_limit=100,
            total_limit=800,
            status="pending",
        )
        db_session.add(pending)
        db_session.commit()

        rows = [make_valid_row(
            place_url="https://m.place.naver.com/restaurant/77777",
        )]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        campaign = data["campaigns"][0]
        assert campaign["extension_eligible"] is False

    def test_preview_multiple_rows_mixed(self, client, db_session):
        """여러 행: 연장 가능 + 연장 불가 혼합."""
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()

        # place_id=11111 -> active 캠페인 존재 (연장 가능)
        existing1 = Campaign(
            campaign_code="C11111",
            account_id=account.id,
            place_name="플레이스1",
            place_url="https://m.place.naver.com/restaurant/11111",
            place_id="11111",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=7),
            end_date=date.today() + timedelta(days=7),
            daily_limit=100,
            total_limit=2000,
            status="active",
        )
        db_session.add(existing1)
        db_session.commit()

        rows = [
            # 행1: place_id=11111 -> 연장 가능
            make_valid_row(place_url="https://m.place.naver.com/restaurant/11111"),
            # 행2: place_id=22222 -> 기존 없음, 연장 불가
            make_valid_row(place_url="https://m.place.naver.com/restaurant/22222"),
        ]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["total_count"] == 2

        # 행1: 연장 가능
        assert data["campaigns"][0]["extension_eligible"] is True
        assert data["campaigns"][0]["existing_campaign_code"] == "C11111"
        # 행2: 연장 불가
        assert data["campaigns"][1]["extension_eligible"] is False
        assert data["campaigns"][1]["existing_campaign_code"] is None

    def test_preview_invalid_row_skips_extension_check(self, client, db_session):
        """유효하지 않은 행은 연장 체크를 건너뜀."""
        rows = [
            [
                "대행사", "testuser1",
                (date.today() + timedelta(days=1)).isoformat(),
                (date.today() + timedelta(days=7)).isoformat(),
                100, "키워드1,키워드2",  # 키워드 10개 미만 -> invalid
                "플레이스",
                "https://m.place.naver.com/restaurant/33333",
                "트래픽",
            ]
        ]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        campaign = data["campaigns"][0]

        assert campaign["is_valid"] is False
        assert campaign["extension_eligible"] is False


# ============================================================
# Confirm API - 연장/신규 분기 처리 테스트
# ============================================================

class TestConfirmExtension:
    """POST /upload/confirm - 연장/신규 분기 처리 테스트."""

    def _base_confirm_item(self, **overrides):
        """기본 확정 캠페인 데이터 생성."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "agency_name": "테스트대행사1",
            "user_id": "testuser1",
            "start_date": tomorrow.isoformat(),
            "end_date": next_week.isoformat(),
            "daily_limit": 100,
            "keywords": [f"키워드{i}" for i in range(15)],
            "place_name": "테스트플레이스",
            "place_url": "https://m.place.naver.com/restaurant/12345",
            "campaign_type": "트래픽",
            "action": "new",
        }
        data.update(overrides)
        return data

    def test_confirm_new_campaign(self, client, db_session):
        """신규 캠페인 확정 - status=pending (자동등록 대기)."""
        data = {"campaigns": [self._base_confirm_item(action="new")]}

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        resp = response.json()
        assert resp["success"] is True
        assert resp["created_count"] == 1
        assert resp["new_count"] == 1
        assert resp["extend_count"] == 0

        # DB 확인
        campaign = db_session.query(Campaign).first()
        assert campaign.status == "pending"
        assert campaign.registration_step == "queued"
        assert campaign.extend_target_id is None
        assert campaign.place_id == "12345"

    def test_confirm_extend_campaign(self, client, db_session):
        """연장 캠페인 확정 - status=pending_extend."""
        # 기존 active 캠페인 생성
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        existing = Campaign(
            campaign_code="CEXT1",
            account_id=account.id,
            place_name="테스트플레이스",
            place_url="https://m.place.naver.com/restaurant/12345",
            place_id="12345",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=7),
            end_date=date.today() + timedelta(days=7),
            daily_limit=100,
            total_limit=1400,
            status="active",
        )
        db_session.add(existing)
        db_session.commit()

        data = {"campaigns": [self._base_confirm_item(
            action="extend",
            existing_campaign_id=existing.id,
        )]}

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        resp = response.json()
        assert resp["success"] is True
        assert resp["created_count"] == 1
        assert resp["new_count"] == 0
        assert resp["extend_count"] == 1

        # DB 확인 - 새로 생성된 pending_extend 캠페인
        new_campaign = db_session.query(Campaign).filter(
            Campaign.status == "pending_extend"
        ).first()
        assert new_campaign is not None
        assert new_campaign.extend_target_id == existing.id
        assert new_campaign.place_id == "12345"

    def test_confirm_extend_without_existing_campaign_id(self, client, db_session):
        """existing_campaign_id 없이 extend 요청 시 신규로 처리."""
        data = {"campaigns": [self._base_confirm_item(
            action="extend",
            # existing_campaign_id 없음
        )]}

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        resp = response.json()
        # existing_campaign_id가 None이므로 else 분기 -> 신규로 처리
        assert resp["created_count"] == 1
        assert resp["new_count"] == 1
        assert resp["extend_count"] == 0

    def test_confirm_extend_nonexistent_target(self, client, db_session):
        """존재하지 않는 대상 캠페인으로 extend 시 건너뛰기."""
        data = {"campaigns": [self._base_confirm_item(
            action="extend",
            existing_campaign_id=99999,  # 존재하지 않는 ID
        )]}

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        resp = response.json()
        assert resp["created_count"] == 0
        assert resp["extend_count"] == 0

    def test_confirm_extend_inactive_target(self, client, db_session):
        """대상 캠페인이 active 아니면 건너뛰기."""
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        completed = Campaign(
            campaign_code="CDONE",
            account_id=account.id,
            place_name="완료된플레이스",
            place_url="https://m.place.naver.com/restaurant/55555",
            place_id="55555",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=14),
            end_date=date.today() - timedelta(days=7),
            daily_limit=100,
            total_limit=800,
            status="completed",
        )
        db_session.add(completed)
        db_session.commit()

        data = {"campaigns": [self._base_confirm_item(
            action="extend",
            existing_campaign_id=completed.id,
        )]}

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        resp = response.json()
        assert resp["created_count"] == 0
        assert resp["extend_count"] == 0

    def test_confirm_mixed_new_and_extend(self, client, db_session):
        """신규 + 연장 혼합 요청."""
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        existing = Campaign(
            campaign_code="CMIX1",
            account_id=account.id,
            place_name="기존플레이스",
            place_url="https://m.place.naver.com/restaurant/44444",
            place_id="44444",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=7),
            end_date=date.today() + timedelta(days=7),
            daily_limit=100,
            total_limit=1400,
            status="active",
        )
        db_session.add(existing)
        db_session.commit()

        data = {"campaigns": [
            # 신규
            self._base_confirm_item(
                place_name="새플레이스",
                place_url="https://m.place.naver.com/restaurant/88888",
                action="new",
            ),
            # 연장
            self._base_confirm_item(
                place_name="기존플레이스",
                place_url="https://m.place.naver.com/restaurant/44444",
                action="extend",
                existing_campaign_id=existing.id,
            ),
        ]}

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        resp = response.json()
        assert resp["created_count"] == 2
        assert resp["new_count"] == 1
        assert resp["extend_count"] == 1

        # DB 확인 - 신규 캠페인은 pending 상태 (자동등록 대기)
        new_campaign = db_session.query(Campaign).filter(Campaign.status == "pending", Campaign.place_id == "88888").first()
        assert new_campaign is not None
        assert new_campaign.registration_step == "queued"

        pending_extend = db_session.query(Campaign).filter(Campaign.status == "pending_extend").first()
        assert pending_extend is not None
        assert pending_extend.extend_target_id == existing.id
        assert pending_extend.place_id == "44444"

    def test_confirm_message_new_only(self, client, db_session):
        """신규만 있을 때 메시지 확인."""
        data = {"campaigns": [self._base_confirm_item(action="new")]}

        response = client.post("/upload/confirm", json=data)
        resp = response.json()
        assert "신규 1개" in resp["message"]
        assert "연장" not in resp["message"]

    def test_confirm_message_extend_only(self, client, db_session):
        """연장만 있을 때 메시지 확인."""
        account = db_session.query(Account).filter(Account.user_id == "testuser1").first()
        existing = Campaign(
            campaign_code="CMSG",
            account_id=account.id,
            place_name="플레이스",
            place_url="https://m.place.naver.com/restaurant/66666",
            place_id="66666",
            campaign_type="트래픽",
            start_date=date.today() - timedelta(days=7),
            end_date=date.today() + timedelta(days=7),
            daily_limit=100,
            total_limit=1400,
            status="active",
        )
        db_session.add(existing)
        db_session.commit()

        data = {"campaigns": [self._base_confirm_item(
            place_url="https://m.place.naver.com/restaurant/66666",
            action="extend",
            existing_campaign_id=existing.id,
        )]}

        response = client.post("/upload/confirm", json=data)
        resp = response.json()
        assert "연장 1개" in resp["message"]
        assert "신규" not in resp["message"]


# ============================================================
# 스키마 유효성 검증 테스트
# ============================================================

class TestSchemaValidation:
    """Pydantic 스키마 유효성 검증 테스트."""

    def test_confirm_invalid_action(self, client, db_session):
        """잘못된 action 값 검증."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "campaigns": [{
                "agency_name": "대행사",
                "user_id": "testuser1",
                "start_date": tomorrow.isoformat(),
                "end_date": next_week.isoformat(),
                "daily_limit": 100,
                "keywords": ["키워드"],
                "place_name": "플레이스",
                "place_url": "https://m.place.naver.com/place/1",
                "campaign_type": "트래픽",
                "action": "invalid_action",
            }]
        }

        response = client.post("/upload/confirm", json=data)
        assert response.status_code == 422  # Pydantic validation error

    def test_confirm_default_action_is_new(self, client, db_session):
        """action 기본값은 'new'."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "campaigns": [{
                "agency_name": "대행사",
                "user_id": "testuser1",
                "start_date": tomorrow.isoformat(),
                "end_date": next_week.isoformat(),
                "daily_limit": 100,
                "keywords": [f"키워드{i}" for i in range(15)],
                "place_name": "플레이스",
                "place_url": "https://m.place.naver.com/restaurant/12345",
                "campaign_type": "트래픽",
                # action 미지정 -> 기본값 "new"
            }]
        }

        response = client.post("/upload/confirm", json=data)
        assert response.status_code == 200
        resp = response.json()
        assert resp["new_count"] == 1
        assert resp["extend_count"] == 0

    def test_confirm_place_id_extracted(self, client, db_session):
        """확정 시 place_id가 URL에서 자동 추출되는지 확인."""
        data = {
            "campaigns": [{
                "agency_name": "대행사",
                "user_id": "testuser1",
                "start_date": (date.today() + timedelta(days=1)).isoformat(),
                "end_date": (date.today() + timedelta(days=7)).isoformat(),
                "daily_limit": 100,
                "keywords": [f"키워드{i}" for i in range(15)],
                "place_name": "플레이스",
                "place_url": "https://m.place.naver.com/restaurant/9876543",
                "campaign_type": "트래픽",
                "action": "new",
            }]
        }

        response = client.post("/upload/confirm", json=data)
        assert response.status_code == 200

        campaign = db_session.query(Campaign).first()
        assert campaign.place_id == "9876543"


# ============================================================
# 기존 테스트 하위 호환성
# ============================================================

class TestBackwardCompatibility:
    """기존 API와의 하위 호환성 테스트."""

    def test_preview_response_includes_extension_fields(self, client, db_session):
        """미리보기 응답에 연장 필드가 항상 포함."""
        rows = [make_valid_row()]
        excel_bytes = create_test_excel(rows)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        data = response.json()
        campaign = data["campaigns"][0]

        # 연장 필드가 존재하는지 확인
        assert "extension_eligible" in campaign
        assert "existing_campaign_code" in campaign
        assert "existing_campaign_id" in campaign
        assert "existing_total_count" in campaign

    def test_confirm_without_action_field(self, client, db_session):
        """action 필드 없이 요청해도 기본값 new로 동작."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "campaigns": [{
                "agency_name": "대행사",
                "user_id": "testuser1",
                "start_date": tomorrow.isoformat(),
                "end_date": next_week.isoformat(),
                "daily_limit": 100,
                "keywords": [f"키워드{i}" for i in range(15)],
                "place_name": "플레이스",
                "place_url": "https://m.place.naver.com/restaurant/12345",
                "campaign_type": "트래픽",
            }]
        }

        response = client.post("/upload/confirm", json=data)
        assert response.status_code == 200
        resp = response.json()
        assert resp["success"] is True
        assert resp["new_count"] == 1

    def test_confirm_response_includes_counts(self, client, db_session):
        """응답에 new_count, extend_count 필드가 포함."""
        data = {
            "campaigns": [{
                "agency_name": "대행사",
                "user_id": "testuser1",
                "start_date": (date.today() + timedelta(days=1)).isoformat(),
                "end_date": (date.today() + timedelta(days=7)).isoformat(),
                "daily_limit": 100,
                "keywords": [f"키워드{i}" for i in range(15)],
                "place_name": "플레이스",
                "place_url": "https://m.place.naver.com/restaurant/12345",
                "campaign_type": "트래픽",
            }]
        }

        response = client.post("/upload/confirm", json=data)
        resp = response.json()

        assert "new_count" in resp
        assert "extend_count" in resp
        assert "created_count" in resp
