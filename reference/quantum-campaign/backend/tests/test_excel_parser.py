"""ExcelParser 단위 테스트."""

import atexit
import gc
import os
import tempfile
from datetime import date, timedelta

import pytest
from openpyxl import Workbook

from app.services.excel_parser import (
    ExcelParser,
    CampaignData,
    ParseResult,
    REQUIRED_COLUMNS,
    MIN_KEYWORD_COUNT,
)

# Windows에서 임시 파일 정리를 위한 리스트
_temp_files_to_cleanup = []


def cleanup_temp_files():
    """테스트 종료 시 임시 파일 정리."""
    gc.collect()  # 가비지 컬렉션 강제 실행
    for f in _temp_files_to_cleanup:
        try:
            if os.path.exists(f):
                os.unlink(f)
        except (PermissionError, OSError):
            pass  # Windows에서 여전히 사용 중이면 무시


atexit.register(cleanup_temp_files)


class TestExcelParser:
    """ExcelParser 클래스 테스트."""

    @pytest.fixture
    def parser(self):
        """기본 파서 인스턴스."""
        return ExcelParser()

    @pytest.fixture
    def parser_with_users(self):
        """유효한 사용자 ID가 설정된 파서."""
        return ExcelParser(valid_user_ids=["user1", "user2", "testuser"])

    @pytest.fixture
    def valid_excel_file(self):
        """유효한 데이터가 있는 엑셀 파일 생성."""
        workbook = Workbook()
        sheet = workbook.active

        # 헤더
        headers = REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # 유효한 데이터 행
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        valid_row = [
            "테스트대행사",  # 대행사명
            "testuser",  # 사용자ID
            tomorrow,  # 시작일
            next_week,  # 마감일
            100,  # 일일 한도
            keywords,  # 키워드
            "https://m.place.naver.com/restaurant/12345",  # 플레이스 URL
            "트래픽",  # 캠페인 이름
        ]

        for col, value in enumerate(valid_row, 1):
            sheet.cell(row=2, column=col, value=value)

        # 임시 파일 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()  # 명시적으로 파일 핸들 닫기
        workbook.save(temp_file.name)
        workbook.close()
        _temp_files_to_cleanup.append(temp_file.name)

        yield temp_file.name

        # 정리 시도 (실패해도 atexit에서 재시도)
        gc.collect()
        try:
            os.unlink(temp_file.name)
        except (PermissionError, OSError):
            pass

    @pytest.fixture
    def missing_column_excel_file(self):
        """필수 컬럼이 누락된 엑셀 파일."""
        workbook = Workbook()
        sheet = workbook.active

        # 일부 컬럼만 포함
        headers = ["대행사명", "사용자ID", "시작일"]  # 나머지 컬럼 누락
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        workbook.close()
        _temp_files_to_cleanup.append(temp_file.name)

        yield temp_file.name

        gc.collect()
        try:
            os.unlink(temp_file.name)
        except (PermissionError, OSError):
            pass

    @pytest.fixture
    def invalid_data_excel_file(self):
        """유효하지 않은 데이터가 있는 엑셀 파일."""
        workbook = Workbook()
        sheet = workbook.active

        # 헤더
        for col, header in enumerate(REQUIRED_COLUMNS, 1):
            sheet.cell(row=1, column=col, value=header)

        # 유효하지 않은 데이터 행
        yesterday = date.today() - timedelta(days=1)
        invalid_row = [
            "테스트대행사",  # 대행사명
            "invalid_user",  # 존재하지 않는 사용자ID
            yesterday,  # 과거 시작일
            yesterday - timedelta(days=1),  # 시작일보다 이전인 마감일
            -10,  # 음수 일일 한도
            "키워드1,키워드2",  # 10개 미만 키워드
            "테스트플레이스",  # 플레이스 상호명
            "https://invalid.com/place",  # 잘못된 URL
            "잘못된타입",  # 잘못된 타입구분
        ]

        for col, value in enumerate(invalid_row, 1):
            sheet.cell(row=2, column=col, value=value)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        workbook.close()
        _temp_files_to_cleanup.append(temp_file.name)

        yield temp_file.name

        gc.collect()
        try:
            os.unlink(temp_file.name)
        except (PermissionError, OSError):
            pass

    # === 파싱 테스트 ===

    def test_parse_valid_excel(self, parser_with_users, valid_excel_file):
        """정상 엑셀 파싱 테스트."""
        result = parser_with_users.parse(valid_excel_file)

        assert result.success is True
        assert len(result.campaigns) == 1
        assert len(result.errors) == 0

        campaign = result.campaigns[0]
        assert campaign.is_valid is True
        assert campaign.agency_name == "테스트대행사"
        assert campaign.user_id == "testuser"
        assert campaign.daily_limit == 100
        assert len(campaign.keywords) == 15
        assert campaign.campaign_type == "트래픽"

    def test_parse_missing_column(self, parser, missing_column_excel_file):
        """필수 컬럼 누락 시 에러 테스트."""
        result = parser.parse(missing_column_excel_file)

        assert result.success is False
        assert len(result.errors) > 0
        assert "필수 컬럼이 누락되었습니다" in result.errors[0]

    def test_parse_file_not_found(self, parser):
        """존재하지 않는 파일 테스트."""
        result = parser.parse("/nonexistent/path/file.xlsx")

        assert result.success is False
        assert len(result.errors) > 0
        assert "파일을 찾을 수 없습니다" in result.errors[0]

    def test_parse_invalid_data(self, parser_with_users, invalid_data_excel_file):
        """유효하지 않은 데이터 검증 테스트."""
        result = parser_with_users.parse(invalid_data_excel_file)

        # 파싱은 성공하지만 캠페인 데이터에 에러가 있음
        assert len(result.campaigns) == 1
        assert result.success is False  # 에러가 있으므로 실패

        campaign = result.campaigns[0]
        assert campaign.is_valid is False
        assert len(campaign.errors) > 0

        # 예상되는 에러들 확인
        error_text = " ".join(campaign.errors)
        assert "사용자ID" in error_text or "존재하지 않는" in error_text
        assert "시작일" in error_text or "오늘 이후" in error_text
        assert "일일 한도" in error_text or "0보다" in error_text
        assert "키워드" in error_text or "최소" in error_text
        assert "URL" in error_text

    # === URL 검증 테스트 ===

    def test_validate_url_valid(self, parser):
        """유효한 URL 테스트."""
        valid_urls = [
            "https://m.place.naver.com/restaurant/12345",
            "https://m.place.naver.com/place/67890",
            "http://m.place.naver.com/cafe/11111",
        ]

        for url in valid_urls:
            assert parser.validate_url(url) is True

    def test_validate_url_invalid(self, parser):
        """유효하지 않은 URL 테스트."""
        invalid_urls = [
            "https://place.naver.com/restaurant/12345",  # m. 없음
            "https://m.map.naver.com/place/12345",  # 다른 도메인
            "https://google.com/maps",
            "",
            None,
        ]

        for url in invalid_urls:
            assert parser.validate_url(url) is False

    # === 키워드 파싱 테스트 ===

    def test_parse_keywords_normal(self, parser):
        """정상 키워드 파싱 테스트."""
        keywords_str = "맛집,카페,음식점,디저트,브런치"
        keywords = parser.parse_keywords(keywords_str)

        assert len(keywords) == 5
        assert keywords == ["맛집", "카페", "음식점", "디저트", "브런치"]

    def test_parse_keywords_with_spaces(self, parser):
        """공백이 포함된 키워드 파싱 테스트."""
        keywords_str = " 맛집 , 카페 ,  음식점  "
        keywords = parser.parse_keywords(keywords_str)

        assert len(keywords) == 3
        assert keywords == ["맛집", "카페", "음식점"]

    def test_parse_keywords_empty(self, parser):
        """빈 키워드 테스트."""
        assert parser.parse_keywords("") == []
        assert parser.parse_keywords(None) == []

    def test_parse_keywords_with_empty_items(self, parser):
        """빈 항목이 있는 키워드 파싱 테스트."""
        keywords_str = "맛집,,카페,  ,음식점"
        keywords = parser.parse_keywords(keywords_str)

        assert len(keywords) == 3
        assert keywords == ["맛집", "카페", "음식점"]

    def test_parse_keywords_count_validation(self, parser_with_users, valid_excel_file):
        """키워드 개수 검증 테스트."""
        # valid_excel_file은 15개 키워드를 포함
        result = parser_with_users.parse(valid_excel_file)
        campaign = result.campaigns[0]

        assert len(campaign.keywords) >= MIN_KEYWORD_COUNT
        assert campaign.is_valid is True

    # === 날짜 검증 테스트 ===

    def test_validate_date_formats(self, parser):
        """다양한 날짜 형식 파싱 테스트."""
        # _parse_date는 private 메서드이지만 테스트 목적으로 접근
        assert parser._parse_date("2025-12-25") == date(2025, 12, 25)
        assert parser._parse_date("2025/12/25") == date(2025, 12, 25)
        assert parser._parse_date("2025.12.25") == date(2025, 12, 25)

    def test_validate_date_invalid(self, parser):
        """유효하지 않은 날짜 테스트."""
        assert parser._parse_date("invalid") is None
        assert parser._parse_date("") is None
        assert parser._parse_date(None) is None

    # === 데이터클래스 테스트 ===

    def test_campaign_data_is_valid(self):
        """CampaignData.is_valid 프로퍼티 테스트."""
        valid_campaign = CampaignData(
            agency_name="test",
            user_id="user1",
            start_date=date.today(),
            end_date=date.today(),
            daily_limit=100,
            keywords=["kw1"],
            place_name="place",
            place_url="url",
            campaign_type="트래픽",
            errors=[]
        )
        assert valid_campaign.is_valid is True

        invalid_campaign = CampaignData(
            agency_name="test",
            user_id="user1",
            start_date=date.today(),
            end_date=date.today(),
            daily_limit=100,
            keywords=["kw1"],
            place_name="place",
            place_url="url",
            campaign_type="트래픽",
            errors=["에러가 있습니다"]
        )
        assert invalid_campaign.is_valid is False

    def test_parse_result_valid_invalid_campaigns(self):
        """ParseResult의 valid/invalid campaigns 프로퍼티 테스트."""
        campaigns = [
            CampaignData(
                agency_name="test1",
                user_id="user1",
                start_date=date.today(),
                end_date=date.today(),
                daily_limit=100,
                keywords=["kw1"],
                place_name="place1",
                place_url="url1",
                campaign_type="트래픽",
                errors=[]
            ),
            CampaignData(
                agency_name="test2",
                user_id="user2",
                start_date=date.today(),
                end_date=date.today(),
                daily_limit=100,
                keywords=["kw2"],
                place_name="place2",
                place_url="url2",
                campaign_type="저장하기",
                errors=["에러"]
            ),
        ]

        result = ParseResult(success=False, campaigns=campaigns, errors=[])

        assert len(result.valid_campaigns) == 1
        assert len(result.invalid_campaigns) == 1
        assert result.valid_campaigns[0].agency_name == "test1"
        assert result.invalid_campaigns[0].agency_name == "test2"


class TestExcelParserMultipleRows:
    """다중 행 파싱 테스트."""

    @pytest.fixture
    def multi_row_excel_file(self):
        """여러 행이 있는 엑셀 파일."""
        workbook = Workbook()
        sheet = workbook.active

        # 헤더
        for col, header in enumerate(REQUIRED_COLUMNS, 1):
            sheet.cell(row=1, column=col, value=header)

        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        # 3개의 유효한 행
        for row_num in range(2, 5):
            row_data = [
                f"대행사{row_num}",
                f"user{row_num}",
                tomorrow,
                next_week,
                100 * row_num,
                keywords,
                f"플레이스{row_num}",
                f"https://m.place.naver.com/place/{row_num}",
                "트래픽" if row_num % 2 == 0 else "저장하기",
            ]
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col, value=value)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        workbook.close()
        _temp_files_to_cleanup.append(temp_file.name)

        yield temp_file.name

        gc.collect()
        try:
            os.unlink(temp_file.name)
        except (PermissionError, OSError):
            pass

    def test_parse_multiple_rows(self, multi_row_excel_file):
        """다중 행 파싱 테스트."""
        parser = ExcelParser(valid_user_ids=["user2", "user3", "user4"])
        result = parser.parse(multi_row_excel_file)

        assert len(result.campaigns) == 3

        # 각 행 데이터 확인
        for i, campaign in enumerate(result.campaigns, start=2):
            assert campaign.agency_name == f"대행사{i}"
            assert campaign.daily_limit == 100 * i
            assert campaign.row_number == i

    def test_skip_empty_rows(self):
        """빈 행 건너뛰기 테스트."""
        workbook = Workbook()
        sheet = workbook.active

        # 헤더
        for col, header in enumerate(REQUIRED_COLUMNS, 1):
            sheet.cell(row=1, column=col, value=header)

        # 2행: 유효한 데이터
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        row_data = [
            "대행사", "testuser", tomorrow, next_week, 100,
            keywords, "플레이스", "https://m.place.naver.com/place/1", "트래픽"
        ]
        for col, value in enumerate(row_data, 1):
            sheet.cell(row=2, column=col, value=value)

        # 3행: 빈 행
        # 4행: 유효한 데이터
        for col, value in enumerate(row_data, 1):
            sheet.cell(row=4, column=col, value=value)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        workbook.close()
        _temp_files_to_cleanup.append(temp_file.name)

        try:
            parser = ExcelParser(valid_user_ids=["testuser"])
            result = parser.parse(temp_file.name)

            # 빈 행은 건너뛰어야 함
            assert len(result.campaigns) == 2
        finally:
            gc.collect()
            try:
                os.unlink(temp_file.name)
            except (PermissionError, OSError):
                pass
