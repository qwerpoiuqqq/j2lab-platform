"""Upload API 통합 테스트."""

import io
import os
import tempfile
from datetime import date, timedelta

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.main import app
from app.models.account import Account
from app.models.campaign import Campaign
from app.services.excel_parser import REQUIRED_COLUMNS


# 테스트용 인메모리 DB 설정
SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"

engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture(scope="function")
def db_session():
    """테스트용 DB 세션."""
    Base.metadata.create_all(bind=engine)
    session = TestingSessionLocal()

    # 테스트용 계정 추가
    test_accounts = [
        Account(user_id="testuser1", agency_name="테스트대행사1", is_active=True),
        Account(user_id="testuser2", agency_name="테스트대행사2", is_active=True),
        Account(user_id="inactive_user", agency_name="비활성대행사", is_active=False),
    ]
    for account in test_accounts:
        session.add(account)
    session.commit()

    yield session

    session.close()
    Base.metadata.drop_all(bind=engine)


@pytest.fixture(scope="function")
def client(db_session):
    """테스트 클라이언트."""
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def create_test_excel(rows_data: list) -> bytes:
    """테스트용 엑셀 파일 생성."""
    workbook = Workbook()
    sheet = workbook.active

    # 헤더
    for col, header in enumerate(REQUIRED_COLUMNS, 1):
        sheet.cell(row=1, column=col, value=header)

    # 데이터 행
    for row_num, row_data in enumerate(rows_data, start=2):
        for col, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col, value=value)

    # 바이트로 변환
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


class TestUploadPreviewAPI:
    """POST /upload/preview API 테스트."""

    @pytest.fixture
    def valid_excel_data(self):
        """유효한 엑셀 데이터."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        return [
            [
                "테스트대행사1",  # 대행사명
                "testuser1",  # 사용자ID
                tomorrow.isoformat(),  # 시작일
                next_week.isoformat(),  # 마감일
                100,  # 일일 한도
                keywords,  # 키워드
                "https://m.place.naver.com/restaurant/12345",  # 플레이스 URL
                "트래픽",  # 캠페인 이름
            ]
        ]

    def test_upload_preview_success(self, client, valid_excel_data):
        """정상 업로드 미리보기 테스트."""
        excel_bytes = create_test_excel(valid_excel_data)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()

        assert data["success"] is True
        assert data["total_count"] == 1
        assert data["valid_count"] == 1
        assert data["invalid_count"] == 0
        assert len(data["campaigns"]) == 1
        assert data["campaigns"][0]["agency_name"] == "테스트대행사1"
        assert data["campaigns"][0]["is_valid"] is True

    def test_upload_preview_invalid_file_type(self, client):
        """잘못된 파일 타입 테스트."""
        response = client.post(
            "/upload/preview",
            files={"file": ("test.txt", b"not an excel file", "text/plain")}
        )

        assert response.status_code == 400
        assert "엑셀 파일" in response.json()["detail"]

    def test_upload_preview_missing_columns(self, client):
        """필수 컬럼 누락 테스트."""
        workbook = Workbook()
        sheet = workbook.active
        # 일부 컬럼만 포함
        sheet.cell(row=1, column=1, value="대행사명")
        sheet.cell(row=1, column=2, value="사용자ID")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", output.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is False
        assert len(data["file_errors"]) > 0
        assert "필수 컬럼이 누락" in data["file_errors"][0]

    def test_upload_preview_invalid_user_id(self, client):
        """존재하지 않는 사용자ID 테스트."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        rows_data = [
            [
                "대행사", "nonexistent_user", tomorrow.isoformat(), next_week.isoformat(),
                100, keywords, "플레이스", "https://m.place.naver.com/place/1", "트래픽"
            ]
        ]
        excel_bytes = create_test_excel(rows_data)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is False
        assert data["invalid_count"] == 1
        assert "존재하지 않는 사용자ID" in data["campaigns"][0]["errors"][0]

    def test_upload_preview_invalid_url(self, client):
        """잘못된 URL 형식 테스트."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        rows_data = [
            [
                "대행사", "testuser1", tomorrow.isoformat(), next_week.isoformat(),
                100, keywords, "플레이스", "https://invalid.com/place", "트래픽"
            ]
        ]
        excel_bytes = create_test_excel(rows_data)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is False
        assert "URL" in " ".join(data["campaigns"][0]["errors"])

    def test_upload_preview_invalid_date(self, client):
        """잘못된 날짜 테스트 (시작일 > 마감일)."""
        tomorrow = date.today() + timedelta(days=1)
        yesterday = date.today() - timedelta(days=1)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        rows_data = [
            [
                "대행사", "testuser1", tomorrow.isoformat(), yesterday.isoformat(),  # 마감일이 시작일보다 이전
                100, keywords, "플레이스", "https://m.place.naver.com/place/1", "트래픽"
            ]
        ]
        excel_bytes = create_test_excel(rows_data)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is False
        assert "마감일" in " ".join(data["campaigns"][0]["errors"])

    def test_upload_preview_insufficient_keywords(self, client):
        """키워드 개수 부족 테스트."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = "키워드1,키워드2,키워드3"  # 10개 미만

        rows_data = [
            [
                "대행사", "testuser1", tomorrow.isoformat(), next_week.isoformat(),
                100, keywords, "플레이스", "https://m.place.naver.com/place/1", "트래픽"
            ]
        ]
        excel_bytes = create_test_excel(rows_data)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is False
        assert "키워드는 최소" in " ".join(data["campaigns"][0]["errors"])

    def test_upload_preview_multiple_rows(self, client):
        """다중 행 업로드 테스트."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)
        keywords = ",".join([f"키워드{i}" for i in range(15)])

        rows_data = [
            ["대행사1", "testuser1", tomorrow.isoformat(), next_week.isoformat(),
             100, keywords, "https://m.place.naver.com/restaurant/1", "트래픽"],
            ["대행사2", "testuser2", tomorrow.isoformat(), next_week.isoformat(),
             200, keywords, "https://m.place.naver.com/restaurant/2", "저장하기"],
        ]
        excel_bytes = create_test_excel(rows_data)

        response = client.post(
            "/upload/preview",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["total_count"] == 2
        assert data["valid_count"] == 2


class TestUploadConfirmAPI:
    """POST /upload/confirm API 테스트."""

    @pytest.fixture
    def valid_campaign_data(self):
        """유효한 캠페인 확정 데이터."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        return {
            "campaigns": [
                {
                    "agency_name": "테스트대행사1",
                    "user_id": "testuser1",
                    "start_date": tomorrow.isoformat(),
                    "end_date": next_week.isoformat(),
                    "daily_limit": 100,
                    "keywords": [f"키워드{i}" for i in range(15)],
                    "place_name": "테스트플레이스",
                    "place_url": "https://m.place.naver.com/restaurant/12345",
                    "campaign_type": "트래픽"
                }
            ]
        }

    def test_confirm_upload_success(self, client, db_session, valid_campaign_data):
        """정상 업로드 확정 테스트."""
        response = client.post("/upload/confirm", json=valid_campaign_data)

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["created_count"] == 1

        # DB에 캠페인이 생성되었는지 확인
        campaigns = db_session.query(Campaign).all()
        assert len(campaigns) == 1
        assert campaigns[0].place_name == "테스트플레이스"
        assert campaigns[0].status == "pending"
        assert campaigns[0].registration_step == "queued"

    def test_confirm_upload_empty_list(self, client):
        """빈 캠페인 목록 테스트."""
        response = client.post("/upload/confirm", json={"campaigns": []})

        assert response.status_code == 400
        assert "등록할 캠페인이 없습니다" in response.json()["detail"]

    def test_confirm_upload_invalid_user(self, client, db_session):
        """존재하지 않는 사용자 ID 테스트."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "campaigns": [
                {
                    "agency_name": "대행사",
                    "user_id": "nonexistent_user",
                    "start_date": tomorrow.isoformat(),
                    "end_date": next_week.isoformat(),
                    "daily_limit": 100,
                    "keywords": ["키워드"],
                    "place_name": "플레이스",
                    "place_url": "https://m.place.naver.com/place/1",
                    "campaign_type": "트래픽"
                }
            ]
        }

        response = client.post("/upload/confirm", json=data)

        # 존재하지 않는 계정은 건너뛰므로 성공하지만 생성된 것은 없음
        assert response.status_code == 200
        assert response.json()["created_count"] == 0

    def test_confirm_upload_multiple_campaigns(self, client, db_session):
        """다중 캠페인 확정 테스트."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "campaigns": [
                {
                    "agency_name": "대행사1",
                    "user_id": "testuser1",
                    "start_date": tomorrow.isoformat(),
                    "end_date": next_week.isoformat(),
                    "daily_limit": 100,
                    "keywords": ["키워드1"],
                    "place_name": "플레이스1",
                    "place_url": "https://m.place.naver.com/place/1",
                    "campaign_type": "트래픽"
                },
                {
                    "agency_name": "대행사2",
                    "user_id": "testuser2",
                    "start_date": tomorrow.isoformat(),
                    "end_date": next_week.isoformat(),
                    "daily_limit": 200,
                    "keywords": ["키워드2"],
                    "place_name": "플레이스2",
                    "place_url": "https://m.place.naver.com/place/2",
                    "campaign_type": "저장하기"
                }
            ]
        }

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 200
        assert response.json()["created_count"] == 2

        # DB 확인
        campaigns = db_session.query(Campaign).all()
        assert len(campaigns) == 2

    def test_confirm_upload_empty_campaign_type(self, client):
        """빈 캠페인 타입 테스트 (빈 문자열은 거부)."""
        tomorrow = date.today() + timedelta(days=1)
        next_week = date.today() + timedelta(days=7)

        data = {
            "campaigns": [
                {
                    "agency_name": "대행사",
                    "user_id": "testuser1",
                    "start_date": tomorrow.isoformat(),
                    "end_date": next_week.isoformat(),
                    "daily_limit": 100,
                    "keywords": ["키워드"],
                    "place_name": "플레이스",
                    "place_url": "https://m.place.naver.com/place/1",
                    "campaign_type": "  "
                }
            ]
        }

        response = client.post("/upload/confirm", json=data)

        assert response.status_code == 422  # Pydantic validation error
