import json
import logging
import os
from datetime import date, datetime as dt, timedelta as td
from decimal import Decimal
from io import BytesIO
from secrets import compare_digest
from urllib.parse import quote

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from accounts.models import User
from dashboard.models import Notification
from products.models import Category, PricePolicy, Product

from .models import Order
from .services import cancel_order, confirm_payment, create_order
from .validators import validate_order_data

logger = logging.getLogger(__name__)


def _notify_order_status(order):
    Notification.objects.create(
        user=order.user,
        message=f'주문 {order.order_number} 상태: {order.get_status_display()}',
        link=f'/orders/{order.pk}/',
    )


def _safe_excel_text(value):
    text = '' if value is None else str(value)
    if text.startswith(('=', '+', '-', '@')):
        return "'" + text
    return text


@login_required
def order_grid(request):
    if request.user.is_manager:
        return redirect('orders:order_list')
    products = Product.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True).prefetch_related(
        models.Prefetch('products', queryset=Product.objects.filter(is_active=True))
    )
    return render(request, 'orders/order_grid.html', {
        'products': products,
        'categories': categories,
    })


@login_required
@require_POST
def api_order_submit(request):
    if request.user.is_manager:
        return JsonResponse({'success': False, 'errors': [{'row': 0, 'message': '책임자는 접수할 수 없습니다.'}]}, status=403)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'errors': [{'row': 0, 'message': '잘못된 JSON 형식입니다.'}]}, status=400)

    try:
        product_id = body.get('product_id')
        rows = body.get('rows', [])
        memo = body.get('memo', '')

        product = get_object_or_404(Product, pk=product_id, is_active=True)
        valid_rows, errors = validate_order_data(rows, product.schema)
        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)

        if not valid_rows:
            return JsonResponse({'success': False, 'errors': [{'row': 0, 'message': '유효한 데이터가 없습니다.'}]}, status=400)

        order = create_order(request.user, product, valid_rows, memo)
        total = int(order.total_amount)
        supply = int(round(total / Decimal('1.1')))
        vat = total - supply

        return JsonResponse({
            'success': True,
            'order_number': order.order_number,
            'item_count': order.item_count,
            'total_amount': total,
            'supply_amount': supply,
            'vat_amount': vat,
        })
    except ValueError as exc:
        return JsonResponse({'success': False, 'errors': [{'row': 0, 'message': str(exc)}]}, status=400)


@login_required
def api_excel_template_download(request, product_id):
    product = get_object_or_404(Product, pk=product_id, is_active=True)
    schema = product.schema or []

    input_schema = [
        field for field in schema
        if field.get('type') not in ('readonly', 'calc', 'date_calc')
        and not (field.get('sample') and field.get('type') != 'date')
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '주문 데이터'

    for col_idx, field in enumerate(input_schema, 1):
        label = field.get('label', field['name'])
        if field.get('required'):
            label += ' *'
        cell = ws.cell(row=1, column=col_idx, value=label)

        color = (field.get('color') or '#4472C4').lstrip('#')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')

        ftype = field.get('type', 'text')
        if ftype == 'date':
            width = 14
        elif ftype == 'number':
            width = 12
        elif ftype == 'select':
            width = 14
        elif ftype == 'url':
            width = 30
        else:
            width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        if ftype == 'select' and field.get('options'):
            options_str = ','.join(str(opt) for opt in field['options'])
            dv = DataValidation(type='list', formula1=f'"{options_str}"', allow_blank=True)
            dv.error = f'{field.get("label", field["name"])} 값을 목록에서 선택하세요.'
            dv.errorTitle = '입력 오류'
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            dv.add(f'{col_letter}2:{col_letter}1000')
            ws.add_data_validation(dv)

    ws2 = wb.create_sheet('입력 안내')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 40

    type_names = {'text': '텍스트', 'url': 'URL', 'number': '숫자', 'date': '날짜', 'select': '선택'}
    guide_headers = ['필드명', '타입', '필수', '설명']
    for col_idx, header in enumerate(guide_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='F2F4F6', end_color='F2F4F6', fill_type='solid')

    for row_idx, field in enumerate(input_schema, 2):
        ws2.cell(row=row_idx, column=1, value=field.get('label', field['name']))
        ws2.cell(row=row_idx, column=2, value=type_names.get(field.get('type', 'text'), field.get('type')))
        ws2.cell(row=row_idx, column=3, value='Y' if field.get('required') else 'N')

        desc = ''
        if field.get('type') == 'select' and field.get('options'):
            desc = '선택 가능 값: ' + ', '.join(str(opt) for opt in field['options'])
        elif field.get('type') == 'url':
            desc = 'http:// 또는 https:// 로 시작하는 전체 URL'
        elif field.get('type') == 'number':
            desc = '숫자만 입력'
        elif field.get('type') == 'date':
            desc = 'YYYY-MM-DD 형식 (예: 2026-02-15)'

        ws2.cell(row=row_idx, column=4, value=desc)

    auto_fields = [
        field for field in schema
        if field.get('type') in ('calc', 'date_calc') or (field.get('sample') and field.get('type') != 'date')
    ]
    if auto_fields:
        row_idx = len(input_schema) + 3
        ws2.cell(row=row_idx, column=1, value='[자동 처리 필드]').font = Font(bold=True, color='3182F6')
        for field in auto_fields:
            row_idx += 1
            ws2.cell(row=row_idx, column=1, value=field.get('label', field['name']))
            if field.get('type') == 'calc':
                ws2.cell(row=row_idx, column=2, value='자동계산')
            elif field.get('type') == 'date_calc':
                ws2.cell(row=row_idx, column=2, value='자동계산(날짜)')
            else:
                ws2.cell(row=row_idx, column=2, value='고정값')
                ws2.cell(row=row_idx, column=4, value=field.get('sample', ''))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'{product.name}_양식.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


@login_required
@require_POST
def api_excel_upload(request):
    product_id = request.POST.get('product_id')
    file = request.FILES.get('file')

    if not product_id or not file:
        return JsonResponse({'success': False, 'message': '상품과 파일을 선택하세요.'}, status=400)

    max_upload_size = 5 * 1024 * 1024
    if file.size > max_upload_size:
        return JsonResponse({'success': False, 'message': '파일 크기는 5MB 이하여야 합니다.'}, status=400)

    if not file.name.lower().endswith('.xlsx'):
        return JsonResponse({'success': False, 'message': 'xlsx 파일만 업로드할 수 있습니다.'}, status=400)

    ALLOWED_EXCEL_TYPES = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
    }
    if file.content_type not in ALLOWED_EXCEL_TYPES:
        return JsonResponse({'success': False, 'message': '허용되지 않는 파일 형식입니다.'}, status=400)

    product = get_object_or_404(Product, pk=product_id, is_active=True)
    schema = product.schema or []

    input_schema = [
        field for field in schema
        if field.get('type') not in ('readonly', 'calc', 'date_calc')
        and not (field.get('sample') and field.get('type') != 'date')
    ]

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        label_to_name = {}
        for field in input_schema:
            label = field.get('label', field['name'])
            label_to_name[label] = field['name']
            label_to_name[label + ' *'] = field['name']
            label_to_name[field['name']] = field['name']

        col_map = {}
        for col_idx, header in enumerate(header_row):
            key = str(header).strip() if header else ''
            if key in label_to_name:
                col_map[col_idx] = label_to_name[key]

        if not col_map:
            return JsonResponse({'success': False, 'message': '엑셀 헤더가 양식과 일치하지 않습니다.'}, status=400)

        rows = []
        for row in ws.iter_rows(min_row=2):
            row_data = {}
            has_data = False
            for col_idx, name in col_map.items():
                value = row[col_idx].value if col_idx < len(row) else None
                if value is not None:
                    has_data = True
                row_data[name] = str(value).strip() if value is not None else ''
            if not has_data:
                continue

            for field in schema:
                if field.get('sample') and field.get('type') not in ('date', 'calc', 'date_calc'):
                    row_data[field['name']] = field['sample']

            for field in schema:
                if field.get('type') == 'calc' and field.get('formula'):
                    formula = field['formula']
                    try:
                        a = float(row_data.get(formula.get('fieldA', ''), 0) or 0)
                        b = float(row_data.get(formula.get('fieldB', ''), 0) or 0)
                        operator = formula.get('operator', '*')
                        if operator == '*':
                            result = a * b
                        elif operator == '+':
                            result = a + b
                        elif operator == '-':
                            result = a - b
                        elif operator == '/':
                            result = a / b if b != 0 else 0
                        else:
                            result = 0
                        row_data[field['name']] = str(int(result)) if result == int(result) else str(result)
                    except (TypeError, ValueError):
                        row_data[field['name']] = ''

            for field in schema:
                if field.get('type') == 'date_calc' and field.get('formula'):
                    formula = field['formula']
                    try:
                        date_val = row_data.get(formula.get('dateField', ''), '')
                        days = int(float(row_data.get(formula.get('daysField', ''), 0) or 0))
                        if date_val and days > 0:
                            parsed_date = dt.strptime(date_val, '%Y-%m-%d') + td(days=days - 1)
                            row_data[field['name']] = parsed_date.strftime('%Y-%m-%d')
                        else:
                            row_data[field['name']] = ''
                    except (TypeError, ValueError):
                        row_data[field['name']] = ''

            rows.append(row_data)

        wb.close()
        return JsonResponse({'success': True, 'rows': rows, 'count': len(rows)})
    except Exception:
        logger.exception('Excel upload parsing failed')
        return JsonResponse({'success': False, 'message': '엑셀 파일을 처리할 수 없습니다.'}, status=400)


@login_required
def order_list(request):
    user = request.user
    if user.is_admin or user.is_accountant or user.is_manager:
        orders = Order.objects.select_related('user', 'user__parent', 'product', 'approved_by').filter(user_id__in=user.get_all_order_user_ids())
    elif user.is_agency:
        child_ids = User.objects.filter(parent=user).values_list('id', flat=True)
        orders = Order.objects.select_related('user', 'user__parent', 'product', 'approved_by').filter(user_id__in=list(child_ids) + [user.id])
    else:
        orders = Order.objects.select_related('user', 'user__parent', 'product', 'approved_by').filter(user=user)

    status = request.GET.get('status')
    if status:
        orders = orders.filter(status=status)

    q = request.GET.get('q', '').strip()
    if q:
        orders = orders.filter(
            models.Q(order_number__icontains=q) |
            models.Q(user__company_name__icontains=q) |
            models.Q(user__username__icontains=q)
        )

    paginator = Paginator(orders, 20)
    orders_page = paginator.get_page(request.GET.get('page'))
    return render(request, 'orders/order_list.html', {
        'orders': orders_page,
        'status_choices': Order.Status.choices,
        'current_status': status,
        'search_query': q,
    })


@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order.objects.select_related('user', 'user__parent', 'product', 'approved_by'), pk=pk)
    user = request.user
    if user.is_admin or user.is_accountant or user.is_manager:
        if order.user_id not in user.get_all_order_user_ids():
            return redirect('orders:order_list')
    elif user.is_agency:
        child_ids = list(User.objects.filter(parent=user).values_list('id', flat=True))
        if order.user_id not in child_ids + [user.id]:
            return redirect('orders:order_list')
    elif order.user != user:
        return redirect('orders:order_list')

    schema = order.product.schema or []
    columns = [f.get('label', f['name']) for f in schema]
    items = order.items.all()
    item_rows = []
    for item in items:
        values = [item.data.get(f['name'], '') for f in schema]
        item_rows.append({
            'item': item,
            'values': values,
        })

    return render(request, 'orders/order_detail.html', {
        'order': order,
        'items': items,
        'columns': columns,
        'item_rows': item_rows,
    })


@login_required
@require_POST
def order_cancel(request, pk):
    order = get_object_or_404(Order, pk=pk)
    user = request.user
    if not (user.is_admin or user.is_accountant or user.is_manager):
        return redirect('orders:order_list')
    if order.user_id not in user.get_all_order_user_ids():
        return redirect('orders:order_list')

    if order.status not in [Order.Status.SUBMITTED]:
        messages.error(request, '접수완료 상태의 주문만 취소할 수 있습니다.')
        return redirect('orders:order_detail', pk=pk)

    try:
        cancel_order(order, user)
        order.refresh_from_db()
        _notify_order_status(order)
        messages.success(request, '주문이 취소되었습니다.')
    except ValueError as exc:
        messages.error(request, str(exc))
    return redirect('orders:order_detail', pk=pk)


@login_required
@require_POST
def order_delete(request, pk):
    if not (request.user.is_admin or request.user.is_accountant):
        return redirect('orders:order_list')

    order = get_object_or_404(Order, pk=pk)
    if order.user_id not in request.user.get_all_order_user_ids():
        return redirect('orders:order_list')
    order_number = order.order_number
    order.delete()
    messages.success(request, f'주문 {order_number}이(가) 삭제되었습니다.')
    return redirect('orders:order_list')


@login_required
@require_POST
def order_status_update(request, pk):
    if not (request.user.is_admin or request.user.is_accountant or request.user.is_manager):
        return redirect('orders:order_list')

    order = get_object_or_404(Order, pk=pk)
    if order.user_id not in request.user.get_all_order_user_ids():
        return redirect('orders:order_list')
    new_status = request.POST.get('status')
    if new_status in dict(Order.Status.choices):
        order.status = new_status
        order.save(update_fields=['status', 'updated_at'])
        # 주문 항목 상태 동기화
        from .models import OrderItem
        if new_status == Order.Status.PROCESSING:
            order.items.exclude(status=OrderItem.Status.COMPLETED).update(status=OrderItem.Status.PROCESSING)
        elif new_status == Order.Status.COMPLETED:
            order.items.update(status=OrderItem.Status.COMPLETED)
        _notify_order_status(order)
        messages.success(request, f'주문 상태가 {order.get_status_display()}(으)로 변경되었습니다.')
    return redirect('orders:order_detail', pk=pk)


@login_required
@require_POST
def order_bulk_status_update(request):
    if not (request.user.is_admin or request.user.is_accountant):
        return redirect('orders:order_list')

    order_ids = request.POST.getlist('order_ids')
    new_status = request.POST.get('status')
    if not order_ids or new_status not in dict(Order.Status.choices):
        messages.error(request, '주문과 상태를 선택하세요.')
        return redirect('orders:order_list')

    allowed_user_ids = request.user.get_all_order_user_ids()
    orders = Order.objects.filter(pk__in=order_ids, user_id__in=allowed_user_ids)
    from .models import OrderItem
    count = 0
    for order in orders:
        order.status = new_status
        order.save(update_fields=['status', 'updated_at'])
        # 주문 항목 상태 동기화
        if new_status == Order.Status.PROCESSING:
            order.items.exclude(status=OrderItem.Status.COMPLETED).update(status=OrderItem.Status.PROCESSING)
        elif new_status == Order.Status.COMPLETED:
            order.items.update(status=OrderItem.Status.COMPLETED)
        _notify_order_status(order)
        count += 1

    status_label = dict(Order.Status.choices).get(new_status)
    messages.success(request, f'{count}건의 주문이 {status_label}(으)로 변경되었습니다.')
    return redirect('orders:order_list')


@login_required
@require_POST
def order_confirm_payment(request, pk):
    if not (request.user.is_admin or request.user.is_accountant or request.user.is_manager):
        return redirect('orders:order_list')

    order = get_object_or_404(Order, pk=pk)
    if order.user_id not in request.user.get_all_order_user_ids():
        return redirect('orders:order_list')
    try:
        confirm_payment(order, request.user)
        order.refresh_from_db()
        _notify_order_status(order)
        messages.success(request, f'주문 {order.order_number} 입금이 확인되었습니다.')
    except ValueError as exc:
        messages.error(request, str(exc))
    return redirect('orders:order_detail', pk=pk)


@login_required
@require_POST
def order_deadline_update(request, pk):
    if not (request.user.is_admin or request.user.is_accountant or request.user.is_manager):
        return redirect('orders:order_list')

    order = get_object_or_404(Order, pk=pk)
    if order.user_id not in request.user.get_all_order_user_ids():
        return redirect('orders:order_list')
    deadline_str = request.POST.get('deadline')
    if deadline_str:
        try:
            order.deadline = date.fromisoformat(deadline_str)
            order.save(update_fields=['deadline', 'updated_at'])
            messages.success(request, f'마감일이 {order.deadline}(으)로 변경되었습니다.')
        except ValueError:
            messages.error(request, '올바른 날짜 형식이 아닙니다.')
    else:
        order.deadline = None
        order.save(update_fields=['deadline', 'updated_at'])
        messages.info(request, '마감일이 해제되었습니다.')

    return redirect('orders:order_detail', pk=pk)


@login_required
@require_POST
def order_approve(request, pk):
    if not (request.user.is_admin or request.user.is_accountant or request.user.is_manager):
        return redirect('orders:order_list')
    order = get_object_or_404(Order, pk=pk)
    if order.user_id not in request.user.get_all_order_user_ids():
        return redirect('orders:order_list')
    order.approved_by = request.user
    order.approved_at = timezone.now()
    order.save(update_fields=['approved_by', 'approved_at', 'updated_at'])
    approver_name = request.user.first_name or request.user.company_name or request.user.username
    messages.success(request, f'주문이 승인되었습니다. (승인자: {approver_name})')
    return redirect('orders:order_detail', pk=pk)


@login_required
def order_items_export(request, pk):
    """주문 항목 엑셀 다운로드 — 상품 스키마 양식 그대로"""
    order = get_object_or_404(Order.objects.select_related('user', 'user__parent', 'product', 'approved_by'), pk=pk)
    user = request.user

    # 권한 체크
    if user.is_admin or user.is_accountant or user.is_manager:
        if order.user_id not in user.get_all_order_user_ids():
            return redirect('orders:order_list')
    elif user.is_agency:
        child_ids = list(User.objects.filter(parent=user).values_list('id', flat=True))
        if order.user_id not in child_ids + [user.id]:
            return redirect('orders:order_list')
    elif order.user != user:
        return redirect('orders:order_list')

    schema = order.product.schema or []
    items = order.items.all().order_by('row_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '주문 데이터'

    # 헤더 — 스키마 필드 순서 그대로
    header_fill_default = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col_idx, field in enumerate(schema, 1):
        label = field.get('label', field['name'])
        cell = ws.cell(row=1, column=col_idx, value=label)

        color = (field.get('color') or '#4472C4').lstrip('#')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        ftype = field.get('type', 'text')
        if ftype == 'date':
            width = 14
        elif ftype == 'number':
            width = 12
        elif ftype == 'url':
            width = 30
        else:
            width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 데이터 행
    for row_idx, item in enumerate(items, 2):
        for col_idx, field in enumerate(schema, 1):
            value = item.data.get(field['name'], '')
            ws.cell(row=row_idx, column=col_idx, value=_safe_excel_text(value))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    order_date = order.created_at.strftime('%Y-%m-%d')
    company = order.user.company_name or order.user.username
    filename = f'{order_date} - {order.product.name} - {company}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


@login_required
def order_export(request):
    user = request.user
    if user.is_admin or user.is_accountant or user.is_manager:
        orders = Order.objects.select_related('user', 'user__parent', 'product', 'approved_by').filter(user_id__in=user.get_all_order_user_ids())
    elif user.is_agency:
        child_ids = User.objects.filter(parent=user).values_list('id', flat=True)
        orders = Order.objects.select_related('user', 'user__parent', 'product', 'approved_by').filter(user_id__in=list(child_ids) + [user.id])
    else:
        orders = Order.objects.select_related('user', 'user__parent', 'product', 'approved_by').filter(user=user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '주문 목록'

    headers = ['주문번호', '주문자', '상품', '건수', '총 금액', '상태', '주문일']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, order in enumerate(orders, 2):
        ws.cell(row=row_idx, column=1, value=_safe_excel_text(order.order_number))
        ws.cell(row=row_idx, column=2, value=_safe_excel_text(str(order.user)))
        ws.cell(row=row_idx, column=3, value=_safe_excel_text(order.product.name))
        ws.cell(row=row_idx, column=4, value=order.item_count)
        ws.cell(row=row_idx, column=5, value=int(order.total_amount))
        ws.cell(row=row_idx, column=6, value=_safe_excel_text(order.get_status_display()))
        ws.cell(row=row_idx, column=7, value=order.created_at.strftime('%Y-%m-%d %H:%M'))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
    wb.save(response)
    return response


@login_required
def settlement_list(request):
    user = request.user
    if not (user.is_admin or user.is_accountant):
        return redirect('orders:order_list')

    confirmed_statuses = [Order.Status.PAID, Order.Status.PROCESSING, Order.Status.COMPLETED]
    allowed_user_ids = user.get_all_order_user_ids()
    orders = Order.objects.select_related('user', 'product', 'confirmed_by').filter(
        status__in=confirmed_statuses, user_id__in=allowed_user_ids
    )

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    period = request.GET.get('period', 'month')

    if date_from:
        orders = orders.filter(confirmed_at__date__gte=date_from)
    elif period == 'month' and not date_to:
        now = timezone.now()
        first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        orders = orders.filter(confirmed_at__gte=first_day)

    if date_to:
        orders = orders.filter(confirmed_at__date__lte=date_to)

    orders = orders.order_by('-confirmed_at')

    summary = orders.aggregate(total_count=Count('id'), total_amount=Sum('total_amount'))

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '정산 내역'

        headers = ['주문번호', '주문자', '상품', '건수', '금액', '확인일', '확인자', '상태']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1, value=_safe_excel_text(order.order_number))
            ws.cell(row=row_idx, column=2, value=_safe_excel_text(str(order.user)))
            ws.cell(row=row_idx, column=3, value=_safe_excel_text(order.product.name))
            ws.cell(row=row_idx, column=4, value=order.item_count)
            ws.cell(row=row_idx, column=5, value=int(order.total_amount))
            ws.cell(row=row_idx, column=6, value=order.confirmed_at.strftime('%Y-%m-%d %H:%M') if order.confirmed_at else '-')
            ws.cell(row=row_idx, column=7, value=_safe_excel_text(str(order.confirmed_by) if order.confirmed_by else '-'))
            ws.cell(row=row_idx, column=8, value=_safe_excel_text(order.get_status_display()))

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="settlement.xlsx"'
        wb.save(response)
        return response

    orders_page = Paginator(orders, 20).get_page(request.GET.get('page'))

    return render(request, 'orders/settlement_list.html', {
        'orders': orders_page,
        'summary': summary,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'period': period,
    })


@login_required
def api_order_renew_data(request, pk):
    order = get_object_or_404(Order.objects.select_related('user', 'product'), pk=pk)
    user = request.user

    # 권한: 본인 OR admin/accountant/manager/agency(소속 셀러의 주문)
    if user.is_admin or user.is_accountant or user.is_manager:
        if order.user_id not in user.get_all_order_user_ids():
            return JsonResponse({'success': False, 'message': '접근 권한이 없습니다.'}, status=403)
    elif user.is_agency:
        child_ids = list(User.objects.filter(parent=user).values_list('id', flat=True))
        if order.user_id not in child_ids + [user.id]:
            return JsonResponse({'success': False, 'message': '접근 권한이 없습니다.'}, status=403)
    elif order.user != user:
        return JsonResponse({'success': False, 'message': '접근 권한이 없습니다.'}, status=403)

    # 상품이 비활성이면 에러
    if not order.product.is_active:
        return JsonResponse({'success': False, 'message': '해당 상품이 비활성 상태입니다. 재연장할 수 없습니다.'}, status=400)

    rows = list(
        order.items.order_by('row_number').values_list('data', flat=True)
    )

    return JsonResponse({
        'success': True,
        'product_id': order.product_id,
        'product_name': order.product.name,
        'memo': order.memo,
        'rows': rows,
    })


SETTLEMENT_SECRET_PASSWORD = os.getenv('SETTLEMENT_SECRET_PASSWORD', '1019')
SETTLEMENT_SECRET_SESSION_KEY = 'settlement_secret_ok'
SETTLEMENT_SECRET_SESSION_AGE_SECONDS = int(os.getenv('SETTLEMENT_SECRET_SESSION_AGE_SECONDS', '1800'))


@login_required
def settlement_secret(request):
    if not (request.user.is_admin or request.user.is_accountant):
        return redirect('orders:order_list')

    if not SETTLEMENT_SECRET_PASSWORD:
        messages.error(request, '정산 보안 비밀번호가 설정되지 않았습니다.')
        return redirect('orders:settlement_list')

    # 비밀번호 검증
    if request.method == 'POST' and 'password' in request.POST:
        input_password = request.POST.get('password', '')
        if compare_digest(input_password, SETTLEMENT_SECRET_PASSWORD):
            request.session[SETTLEMENT_SECRET_SESSION_KEY] = True
            request.session.set_expiry(SETTLEMENT_SECRET_SESSION_AGE_SECONDS)
        else:
            return render(request, 'orders/settlement_secret_login.html', {'error': True})

    if not request.session.get(SETTLEMENT_SECRET_SESSION_KEY):
        return render(request, 'orders/settlement_secret_login.html')

    confirmed_statuses = [Order.Status.PAID, Order.Status.PROCESSING, Order.Status.COMPLETED]
    allowed_user_ids = request.user.get_all_order_user_ids()
    orders = Order.objects.select_related('user', 'product', 'confirmed_by').filter(
        status__in=confirmed_statuses, user_id__in=allowed_user_ids
    )

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    period = request.GET.get('period', 'month')

    if date_from:
        orders = orders.filter(confirmed_at__date__gte=date_from)
    elif period == 'month' and not date_to:
        now = timezone.now()
        first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        orders = orders.filter(confirmed_at__gte=first_day)

    if date_to:
        orders = orders.filter(confirmed_at__date__lte=date_to)

    orders = orders.order_by('-confirmed_at')

    order_list_all = list(orders)

    # PricePolicy를 미리 한번에 조회 (N+1 쿼리 방지)
    policy_keys = set((o.product_id, o.user_id) for o in order_list_all)
    if policy_keys:
        q = models.Q()
        for pid, uid in policy_keys:
            q |= models.Q(product_id=pid, user_id=uid)
        policies_map = {
            (p.product_id, p.user_id): p
            for p in PricePolicy.objects.filter(q)
        }
    else:
        policies_map = {}

    enriched_orders = []
    sum_total_amount = Decimal('0')
    sum_supply = Decimal('0')
    sum_vat = Decimal('0')
    sum_reduced_qty = 0
    sum_reduced_profit = Decimal('0')

    for order in order_list_all:
        total = int(order.total_amount)
        supply = int(round(Decimal(total) / Decimal('1.1')))
        vat = total - supply

        # 업체별 감은 비율 조회 → 없으면 상품 기본값 사용
        policy = policies_map.get((order.product_id, order.user_id))
        if policy and policy.reduction_rate is not None:
            rate = policy.reduction_rate
        else:
            rate = order.product.reduction_rate or 0
        total_qty = order.total_quantity or 0
        reduced_qty = int(total_qty * rate / 100)
        actual_qty = total_qty - reduced_qty
        per_unit = Decimal(supply) / Decimal(total_qty) if total_qty > 0 else Decimal('0')
        reduced_profit = int(Decimal(reduced_qty) * per_unit)

        enriched_orders.append({
            'order': order,
            'supply': supply,
            'vat': vat,
            'reduction_rate': rate,
            'total_qty': total_qty,
            'reduced_qty': reduced_qty,
            'actual_qty': actual_qty,
            'reduced_profit': reduced_profit,
        })

        sum_total_amount += order.total_amount
        sum_supply += supply
        sum_vat += vat
        sum_reduced_qty += reduced_qty
        sum_reduced_profit += reduced_profit

    summary = {
        'total_count': len(order_list_all),
        'total_amount': int(sum_total_amount),
        'total_supply': int(sum_supply),
        'total_vat': int(sum_vat),
        'total_reduced_qty': sum_reduced_qty,
        'total_reduced_profit': int(sum_reduced_profit),
    }

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '감은 수익 분석'

        headers = ['주문번호', '업체', '상품', '총타수', '감은%', '감은타수',
                   '실투입', '공급가', '부가세', '총액', '감은수익', '승인일', '상태']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, item in enumerate(enriched_orders, 2):
            o = item['order']
            ws.cell(row=row_idx, column=1, value=_safe_excel_text(o.order_number))
            ws.cell(row=row_idx, column=2, value=_safe_excel_text(o.user.company_name or o.user.username))
            ws.cell(row=row_idx, column=3, value=_safe_excel_text(o.product.name))
            ws.cell(row=row_idx, column=4, value=item['total_qty'])
            ws.cell(row=row_idx, column=5, value=item['reduction_rate'])
            ws.cell(row=row_idx, column=6, value=item['reduced_qty'])
            ws.cell(row=row_idx, column=7, value=item['actual_qty'])
            ws.cell(row=row_idx, column=8, value=item['supply'])
            ws.cell(row=row_idx, column=9, value=item['vat'])
            ws.cell(row=row_idx, column=10, value=int(o.total_amount))
            ws.cell(row=row_idx, column=11, value=item['reduced_profit'])
            ws.cell(row=row_idx, column=12, value=o.confirmed_at.strftime('%Y-%m-%d %H:%M') if o.confirmed_at else '-')
            ws.cell(row=row_idx, column=13, value=_safe_excel_text(o.get_status_display()))

        sum_row = len(enriched_orders) + 2
        sum_fill = PatternFill(start_color='F2F4F6', end_color='F2F4F6', fill_type='solid')
        sum_font = Font(bold=True)
        ws.cell(row=sum_row, column=1, value='합계').font = sum_font
        ws.cell(row=sum_row, column=4, value=sum(i['total_qty'] for i in enriched_orders)).font = sum_font
        ws.cell(row=sum_row, column=6, value=summary['total_reduced_qty']).font = sum_font
        ws.cell(row=sum_row, column=7, value=sum(i['actual_qty'] for i in enriched_orders)).font = sum_font
        ws.cell(row=sum_row, column=8, value=summary['total_supply']).font = sum_font
        ws.cell(row=sum_row, column=9, value=summary['total_vat']).font = sum_font
        ws.cell(row=sum_row, column=10, value=summary['total_amount']).font = sum_font
        ws.cell(row=sum_row, column=11, value=summary['total_reduced_profit']).font = sum_font
        for col in range(1, len(headers) + 1):
            ws[f'{chr(64 + col)}{sum_row}'].fill = sum_fill

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 16

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="settlement_secret.xlsx"'
        wb.save(response)
        return response

    paginator = Paginator(enriched_orders, 20)
    orders_page = paginator.get_page(request.GET.get('page'))

    return render(request, 'orders/settlement_secret.html', {
        'orders': orders_page,
        'summary': summary,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'period': period,
    })



