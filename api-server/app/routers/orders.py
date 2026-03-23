"""Orders router: full order lifecycle with state transitions."""

from __future__ import annotations

import io
import logging
from datetime import datetime

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import RoleChecker, get_current_active_user
from app.models.order import Order
from app.models.user import User, UserRole
from app.schemas.common import MessageResponse, PaginatedResponse, PaginationParams
from app.schemas.order import (
    BulkDeleteRequest,
    BulkHoldRequest,
    BulkPaymentConfirmRequest,
    BulkStatusRequest,
    DeadlineUpdateRequest,
    ExcelUploadConfirmRequest,
    ExcelUploadPreviewItem,
    ExcelUploadPreviewResponse,
    OrderBriefResponse,
    OrderCreate,
    OrderHoldRequest,
    OrderRejectRequest,
    OrderResponse,
    OrderUpdate,
    SimplifiedOrderCreate,
)
from app.services import order_service
from app.services import pipeline_orchestrator
from app.services.pipeline_validation import validate_item_data_for_pipeline

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/orders", tags=["orders"])


# ---------------------------------------------------------------------------
# Price masking helper (sub_account must never see pricing data)
# ---------------------------------------------------------------------------

def _mask_prices_for_sub_account(order_data: dict, user_role: str) -> dict:
    """Zero-out all price fields when the requesting user is a sub_account.

    Applied immediately before returning any order response so that the
    masking is enforced regardless of the code path that fetched the data.
    """
    if user_role != UserRole.SUB_ACCOUNT.value:
        return order_data
    order_data["total_amount"] = 0
    order_data["vat_amount"] = 0
    if "items" in order_data and order_data["items"]:
        for item in order_data["items"]:
            item["unit_price"] = 0
            item["subtotal"] = 0
            item["cost_unit_price"] = None
    return order_data


# ---------------------------------------------------------------------------
# Shared helpers (DRY: used by many endpoints)
# ---------------------------------------------------------------------------

async def _get_order_or_404(db: AsyncSession, order_id: int) -> Order:
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    return order


def _check_company_scope(current_user: User, order: Order) -> None:
    """Raise 403 if company_admin tries to access another company's order."""
    if UserRole(current_user.role) == UserRole.COMPANY_ADMIN:
        if order.company_id != current_user.company_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Company admin can only manage orders in their own company",
            )


async def _confirm_and_start_pipeline(
    order_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession,
    current_user: User,
) -> dict:
    """Shared logic for confirm-payment and approve endpoints.

    Pipeline state creation is done synchronously so the response already
    contains the ``processing`` status and pipeline records are visible to
    the frontend immediately.  Only the worker dispatch (HTTP calls to
    keyword-worker) is deferred to a background task.
    """
    order = await _get_order_or_404(db, order_id)
    _check_company_scope(current_user, order)

    try:
        await order_service.confirm_payment(db, order, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    pipeline_warnings: list[str] = []
    confirmed = await order_service.get_order_by_id(db, order_id)
    if confirmed.items:
        for item in confirmed.items:
            for w in validate_item_data_for_pipeline(item.item_data):
                pipeline_warnings.append(f"Item #{item.row_number or item.id}: {w}")

    # Phase 1 (synchronous): create pipeline_state records + set order→processing
    # REQ 6: 입금 확인 시 actor 정보를 타임라인 로그에 기록
    try:
        await pipeline_orchestrator.start_pipeline_for_order(
            db, confirmed,
            actor_id=current_user.id,
            actor_name=current_user.name,
        )
    except Exception:
        logger.exception("Synchronous pipeline start failed for order %s", order_id)

    # Commit before background dispatch to avoid race where the background task
    # runs before pipeline/extraction records become visible to a new session.
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        logger.exception("Commit before extraction dispatch failed for order %s", order_id)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="입금확인 처리 중 커밋 실패",
        )

    # Re-read order after pipeline creation (status is now 'processing')
    confirmed = await order_service.get_order_by_id(db, order_id)

    # Phase 2 (background): dispatch extraction jobs to keyword-worker
    background_tasks.add_task(_dispatch_extraction_jobs_background, order_id)

    result = OrderResponse.model_validate(confirmed).model_dump()
    result["pipeline_warnings"] = pipeline_warnings
    return result


@router.get("/", response_model=PaginatedResponse[OrderBriefResponse])
async def list_orders(
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    status_filter: str | None = Query(None, alias="status"),
    search: str | None = Query(None),
    order_type: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """List orders with role-based filtering and pagination.

    - system_admin: sees all orders
    - company_admin, order_handler: sees orders in the same company
    - distributor: sees own + sub_account orders
    - sub_account: sees only own orders
    """
    pagination = PaginationParams(page=page, size=size)
    orders, total = await order_service.get_orders(
        db,
        skip=pagination.offset,
        limit=pagination.size,
        status=status_filter,
        search=search,
        current_user=current_user,
        order_type=order_type,
    )
    masked_items = [
        _mask_prices_for_sub_account(
            OrderBriefResponse.model_validate(o).model_dump(), current_user.role
        )
        for o in orders
    ]
    return PaginatedResponse.create(
        items=masked_items,
        total=total,
        page=pagination.page,
        size=pagination.size,
    )


@router.post(
    "/",
    response_model=OrderResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_order(
    body: OrderCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Create a new order in draft status.

    Available to all authenticated users (distributor, sub_account typically).
    Prices are resolved automatically based on the user's price tier.
    """
    try:
        order = await order_service.create_order(db, body, current_user)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    # Re-fetch to ensure all nested relationships (items.product) are loaded
    fetched = await order_service.get_order_by_id(db, order.id)
    order_data = OrderResponse.model_validate(fetched).model_dump()
    return _mask_prices_for_sub_account(order_data, current_user.role)


# === Literal-path routes (must be before /{order_id} to avoid path conflicts) ===


@router.post(
    "/simplified",
    response_model=OrderResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_simplified_order(
    body: SimplifiedOrderCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Create a simplified order (no product/category selection needed).

    Automatically matches products by campaign_type, calculates quantities,
    and builds pipeline-compatible item_data.
    """
    try:
        order = await order_service.create_simplified_order(db, body, current_user)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    # Re-fetch to ensure all nested relationships (items.product) are loaded
    fetched = await order_service.get_order_by_id(db, order.id)
    order_data = OrderResponse.model_validate(fetched).model_dump()
    return _mask_prices_for_sub_account(order_data, current_user.role)


@router.get("/excel-template/{product_id}")
async def get_excel_template(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_active_user),
):
    """Generate an Excel template based on a product's form_schema."""
    from openpyxl import Workbook
    from app.services import product_service

    product = await product_service.get_product_by_id(db, product_id)
    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Template"

    # Build headers from form_schema
    headers = ["quantity"]
    if product.form_schema and isinstance(product.form_schema, list):
        for field in product.form_schema:
            if isinstance(field, dict):
                headers.append(field.get("name", field.get("key", "field")))
    elif product.form_schema and isinstance(product.form_schema, dict):
        for key in product.form_schema.get("fields", []):
            if isinstance(key, dict):
                headers.append(key.get("name", key.get("key", "field")))
            elif isinstance(key, str):
                headers.append(key)

    ws.append(headers)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"order_template_{product.id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/excel-upload", response_model=ExcelUploadPreviewResponse)
async def excel_upload(
    file: UploadFile = File(...),
    product_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_active_user),
):
    """Parse an uploaded Excel file, validate against product schema, return preview."""
    from openpyxl import load_workbook
    from app.services import product_service

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are supported",
        )

    product = await product_service.get_product_by_id(db, product_id)
    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found",
        )

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        )

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return ExcelUploadPreviewResponse(
            items=[], total=0, valid_count=0, error_count=0,
            product_id=product_id, product_name=product.name,
        )

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    preview_items = []
    valid_count = 0

    for row_idx, row in enumerate(rows[1:], start=1):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val

        errors = order_service.validate_item_data(product, row_dict)
        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1

        preview_items.append(ExcelUploadPreviewItem(
            row_number=row_idx,
            data=row_dict,
            is_valid=is_valid,
            errors=errors,
        ))

    return ExcelUploadPreviewResponse(
        items=preview_items,
        total=len(preview_items),
        valid_count=valid_count,
        error_count=len(preview_items) - valid_count,
        product_id=product_id,
        product_name=product.name,
    )


@router.post("/excel-upload/confirm", response_model=OrderResponse, status_code=status.HTTP_201_CREATED)
async def excel_upload_confirm(
    body: ExcelUploadConfirmRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Confirm and create order from validated Excel rows."""
    try:
        order = await order_service.create_order_from_excel(
            db, product_id=body.product_id, rows=body.rows,
            current_user=current_user, notes=body.notes,
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    # Re-fetch to ensure all nested relationships (items.product) are loaded
    return await order_service.get_order_by_id(db, order.id)


@router.get("/export")
async def export_orders(
    status_filter: str | None = Query(None, alias="status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER])
    ),
):
    """Export filtered order list as Excel."""
    from openpyxl import Workbook

    orders, total = await order_service.get_orders(
        db, skip=0, limit=50000, status=status_filter, current_user=current_user,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = [
        "ID", "Order Number", "Status", "Payment",
        "Total Amount", "VAT", "Source", "Created At",
    ]
    ws.append(headers)

    for o in orders:
        ws.append([
            o.id,
            o.order_number,
            o.status,
            o.payment_status,
            int(o.total_amount) if o.total_amount else 0,
            int(o.vat_amount) if o.vat_amount else 0,
            o.source,
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orders_export.xlsx"},
    )


@router.post("/bulk-status", response_model=MessageResponse)
async def bulk_status_change(
    body: BulkStatusRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Bulk status change for multiple orders."""
    success_count = 0
    errors = []

    for oid in body.order_ids:
        order = await order_service.get_order_by_id(db, oid)
        if order is None:
            errors.append(f"Order {oid}: not found")
            continue
        try:
            await order_service.transition_order_status(db, order, body.status, current_user)
            success_count += 1
        except ValueError as e:
            errors.append(f"Order {oid}: {str(e)}")

    return MessageResponse(
        message=f"Updated {success_count}/{len(body.order_ids)} orders",
        detail={"errors": errors} if errors else None,
    )


@router.get("/sub-account-pending")
async def get_sub_account_pending_orders(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.DISTRIBUTOR])
    ),
):
    """Get pending sub-account orders awaiting distributor inclusion/exclusion."""
    orders = await order_service.get_sub_account_pending_orders(
        db, distributor=current_user, skip=skip, limit=limit,
    )
    return {
        "items": [
            {
                "id": o.id,
                "order_number": o.order_number,
                "user_id": str(o.user_id),
                "status": o.status,
                "total_amount": int(o.total_amount) if o.total_amount else 0,
                "selection_status": o.selection_status,
                "intake_blocked": order_service.get_order_intake_block_reason(o) is not None,
                "intake_block_reason": order_service.get_order_intake_block_reason(o),
                "created_at": o.created_at.isoformat() if o.created_at else None,
            }
            for o in orders
        ],
    }


@router.get("/distributor-queue")
async def get_distributor_queue(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.DISTRIBUTOR])
    ),
):
    """Get unified queue for a distributor.

    Returns both sub-account pending orders and distributor's own draft/submitted orders.
    Each item includes a 'source' field: 'sub_account' | 'own',
    and 'submitter_name': the name of the order creator.
    """
    entries = await order_service.get_distributor_queue(
        db, distributor=current_user, skip=skip, limit=limit,
    )

    def _serialize(entry: dict) -> dict:
        o = entry["order"]
        source = entry["source"]
        submitter_name = None
        if o.user:
            submitter_name = o.user.name
        items_data = []
        if o.items:
            for item in o.items:
                item_dict = {
                    "product": {"name": item.product.name if item.product else None,
                                "code": item.product.code if item.product else None}
                    if item.product else None,
                    "item_data": item.item_data,
                }
                items_data.append(item_dict)
        return {
            "id": o.id,
            "order_number": o.order_number,
            "user_id": str(o.user_id),
            "user": {"name": o.user.name if o.user else None,
                     "role": o.user.role if o.user else None}
            if o.user else None,
            "status": o.status,
            "total_amount": int(o.total_amount) if o.total_amount else 0,
            "selection_status": o.selection_status,
            "intake_blocked": order_service.get_order_intake_block_reason(o) is not None,
            "intake_block_reason": order_service.get_order_intake_block_reason(o),
            "created_at": o.created_at.isoformat() if o.created_at else None,
            "source": source,
            "submitter_name": submitter_name,
            "items": items_data,
            "item_count": len(o.items) if o.items else 0,
        }

    return {"items": [_serialize(e) for e in entries]}


@router.post("/{order_id}/include")
async def include_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.DISTRIBUTOR])
    ),
):
    """Include a sub-account order (pending -> included)."""
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    try:
        await order_service.set_selection_status(db, order, "included", current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return {"message": "Order included", "order_id": order.id}


@router.post("/{order_id}/exclude")
async def exclude_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.DISTRIBUTOR])
    ),
):
    """Exclude a sub-account order (pending -> excluded) and transition to rejected."""
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    try:
        await order_service.set_selection_status(db, order, "excluded", current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    try:
        from app.services import notification_service
        await notification_service.notify_order_rejected_by_distributor(db, order)
    except Exception:
        pass
    return {"message": "Order excluded", "order_id": order.id, "status": order.status}


@router.post("/bulk-include")
async def bulk_include_orders(
    body: BulkStatusRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.DISTRIBUTOR])
    ),
):
    """Bulk include multiple sub-account orders."""
    success = 0
    errors = []
    for oid in body.order_ids:
        order = await order_service.get_order_by_id(db, oid)
        if order is None:
            errors.append(f"Order {oid}: not found")
            continue
        try:
            await order_service.set_selection_status(db, order, "included", current_user)
            success += 1
        except ValueError as e:
            errors.append(f"Order {oid}: {str(e)}")
    return {"included": success, "errors": errors}


@router.post("/bulk-exclude", response_model=MessageResponse)
async def bulk_exclude_orders(
    body: BulkDeleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.DISTRIBUTOR])
    ),
):
    """Bulk exclude multiple sub-account orders and transition each to rejected."""
    from app.services import notification_service
    success = 0
    errors = []
    for oid in body.order_ids:
        order = await order_service.get_order_by_id(db, oid)
        if order is None:
            errors.append(f"Order {oid}: not found")
            continue
        try:
            await order_service.set_selection_status(db, order, "excluded", current_user)
            success += 1
            try:
                await notification_service.notify_order_rejected_by_distributor(db, order)
            except Exception:
                pass
        except ValueError as e:
            errors.append(f"Order {oid}: {str(e)}")
    return MessageResponse(
        message=f"{success}건 제외 처리되었습니다.",
        detail={"errors": errors} if errors else None,
    )


@router.get("/deadline-batch")
async def get_deadline_batch(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER])
    ),
):
    """마감 처리: 접수 완료(submitted)된 주문 항목을 일괄 확인용으로 반환."""
    from app.models.product import Product

    query = (
        select(Order)
        .where(Order.status.in_(["submitted", "payment_hold"]))
        .order_by(Order.created_at.desc())
    )
    user_role = UserRole(current_user.role)
    if user_role != UserRole.SYSTEM_ADMIN:
        query = query.where(Order.company_id == current_user.company_id)

    result = await db.execute(query)
    orders = list(result.scalars().unique().all())

    items_out = []
    for order in orders:
        for item in (order.items or []):
            product = await db.get(Product, item.product_id)
            items_out.append({
                "order_id": order.id,
                "order_number": order.order_number,
                "user_name": order.user.name if order.user else "-",
                "company_name": order.company.name if order.company else "-",
                "product_id": item.product_id,
                "product_name": product.name if product else "-",
                "daily_deadline": str(product.daily_deadline) if product else "18:00:00",
                "item_id": item.id,
                "item_data": item.item_data or {},
                "quantity": item.quantity,
                "unit_price": int(item.unit_price) if item.unit_price else 0,
                "subtotal": int(item.subtotal) if item.subtotal else 0,
                "status": order.status,
                "created_at": order.created_at.isoformat() if order.created_at else None,
            })
    return {"items": items_out}


@router.get("/deadlines")
async def get_order_deadlines(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get orders with deadlines in given month for calendar view."""
    from datetime import date as d
    from sqlalchemy import func
    from app.models.campaign import Campaign

    start = d(year, month, 1)
    if month == 12:
        end = d(year + 1, 1, 1)
    else:
        end = d(year, month + 1, 1)

    # Orders with deadline in range
    orders_q = await db.execute(
        select(Order).where(
            Order.completed_at.isnot(None),
            func.date(Order.completed_at) >= start,
            func.date(Order.completed_at) < end,
            Order.status.in_(["processing", "payment_confirmed", "completed"]),
        ).order_by(Order.completed_at.asc())
    )
    orders = orders_q.scalars().all()

    # Campaigns with end_date in range
    campaigns_q = await db.execute(
        select(Campaign).where(
            Campaign.end_date >= start,
            Campaign.end_date < end,
            Campaign.status.in_(["active", "completed"]),
        ).order_by(Campaign.end_date.asc())
    )
    campaigns = campaigns_q.scalars().all()

    return {
        "orders": [
            {
                "id": o.id,
                "order_number": o.order_number,
                "deadline": o.completed_at.isoformat() if o.completed_at else None,
                "status": o.status,
                "total_amount": int(o.total_amount) if o.total_amount else 0,
            }
            for o in orders
        ],
        "campaigns": [
            {
                "id": c.id,
                "campaign_code": c.campaign_code,
                "place_name": c.place_name,
                "end_date": c.end_date.isoformat() if c.end_date else None,
                "status": c.status,
            }
            for c in campaigns
        ],
    }


@router.post("/bulk-payment-confirm", response_model=MessageResponse)
async def bulk_payment_confirm(
    body: BulkPaymentConfirmRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER, UserRole.DISTRIBUTOR])
    ),
):
    """Bulk payment confirmation for multiple orders (submitted -> payment_confirmed).

    DISTRIBUTOR: can confirm own orders and sub_account orders they manage.
    """
    success_count = 0
    errors = []
    confirmed_order_ids = []

    for oid in body.order_ids:
        order = await order_service.get_order_by_id(db, oid)
        if order is None:
            errors.append(f"Order {oid}: not found")
            continue

        # company_admin / order_handler can only confirm orders in their company
        user_role = UserRole(current_user.role)
        if user_role in (UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER):
            if order.company_id != current_user.company_id:
                errors.append(f"Order {oid}: not in your company")
                continue

        # distributor can only confirm own or their sub_account orders
        if user_role == UserRole.DISTRIBUTOR:
            if not order_service.can_view_order(current_user, order):
                errors.append(f"Order {oid}: not authorized")
                continue

        try:
            await order_service.confirm_payment(db, order, current_user)
            confirmed_order_ids.append(oid)
            success_count += 1
        except ValueError as e:
            errors.append(f"Order {oid}: {str(e)}")

    # Commit all confirmations before dispatching background pipeline tasks
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="일괄 입금확인 커밋 실패",
        )

    # Start pipelines in BACKGROUND (non-blocking)
    for oid in confirmed_order_ids:
        background_tasks.add_task(_run_pipeline_for_order, oid)

    return MessageResponse(
        message=f"Confirmed {success_count}/{len(body.order_ids)} orders",
        detail={"errors": errors} if errors else None,
    )


@router.post("/bulk-delete", response_model=MessageResponse)
async def bulk_delete_orders(
    body: BulkDeleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Bulk delete multiple orders (only draft/cancelled/rejected)."""
    success_count = 0
    errors = []

    for oid in body.order_ids:
        order = await order_service.get_order_by_id(db, oid)
        if order is None:
            errors.append(f"Order {oid}: not found")
            continue

        # company_admin can only delete orders in their company
        user_role = UserRole(current_user.role)
        if user_role == UserRole.COMPANY_ADMIN:
            if order.company_id != current_user.company_id:
                errors.append(f"Order {oid}: not in your company")
                continue

        try:
            await order_service.delete_order(db, order, current_user)
            success_count += 1
        except ValueError as e:
            errors.append(f"Order {oid}: {str(e)}")

    return MessageResponse(
        message=f"Deleted {success_count}/{len(body.order_ids)} orders",
        detail={"errors": errors} if errors else None,
    )


@router.post("/bulk-hold", response_model=MessageResponse)
async def bulk_hold(
    body: BulkHoldRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Bulk hold for multiple orders."""
    success_count = 0
    errors = []

    for oid in body.order_ids:
        order = await order_service.get_order_by_id(db, oid)
        if order is None:
            errors.append(f"Order {oid}: not found")
            continue

        # company_admin can only hold orders in their company
        user_role = UserRole(current_user.role)
        if user_role == UserRole.COMPANY_ADMIN:
            if order.company_id != current_user.company_id:
                errors.append(f"Order {oid}: not in your company")
                continue

        try:
            await order_service.hold_order(db, order, body.reason, current_user)
            success_count += 1
        except ValueError as e:
            errors.append(f"Order {oid}: {str(e)}")

    return MessageResponse(
        message=f"Held {success_count}/{len(body.order_ids)} orders",
        detail={"errors": errors} if errors else None,
    )


# === Parameterized /{order_id} routes ===


@router.get("/{order_id}", response_model=OrderResponse)
async def get_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get a single order by ID with items."""
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found",
        )

    if not order_service.can_view_order(current_user, order):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this order",
        )

    order_data = OrderResponse.model_validate(order).model_dump()
    return _mask_prices_for_sub_account(order_data, current_user.role)


@router.delete("/{order_id}", response_model=MessageResponse)
async def delete_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Delete an order permanently (only draft/cancelled/rejected)."""
    order = await _get_order_or_404(db, order_id)
    _check_company_scope(current_user, order)

    try:
        await order_service.delete_order(db, order, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return MessageResponse(message=f"Order {order_id} deleted successfully")


@router.patch("/{order_id}", response_model=OrderResponse)
async def update_order(
    order_id: int,
    body: OrderUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker(
            [UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER]
        )
    ),
):
    """Update an order (only in draft status)."""
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found",
        )

    try:
        updated = await order_service.update_order(db, order, notes=body.notes)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    return updated


@router.post("/{order_id}/submit", response_model=OrderResponse)
async def submit_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([
            UserRole.SYSTEM_ADMIN,
            UserRole.COMPANY_ADMIN,
            UserRole.DISTRIBUTOR,
            UserRole.SUB_ACCOUNT,
        ])
    ),
):
    """Submit an order: draft -> submitted.

    - distributor: can submit own + sub_account orders (접수 제출 - 일괄 확인 후 제출)
    - sub_account: can only submit their own orders (제출 요청)
    - system_admin, company_admin: unrestricted
    """
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found",
        )

    user_role = UserRole(current_user.role)

    # Distributors can only submit their own or their sub_account orders
    if user_role == UserRole.DISTRIBUTOR:
        if not order_service.can_view_order(current_user, order):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to submit this order",
            )

    # sub_account can only submit their own orders
    if user_role == UserRole.SUB_ACCOUNT:
        if order.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="하부 대행 계정은 본인의 접수건만 제출 요청할 수 있습니다.",
            )

    try:
        await order_service.submit_order(db, order, current_user)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    return await order_service.get_order_by_id(db, order_id)


@router.post("/{order_id}/confirm-payment", response_model=OrderResponse)
async def confirm_payment(
    order_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER, UserRole.DISTRIBUTOR])
    ),
):
    """Confirm payment: submitted -> payment_confirmed.

    Deducts balance from the order's user.
    - system_admin, company_admin, order_handler: unrestricted (within company scope)
    - distributor: can confirm own orders and sub_account orders they manage
    Pipeline is started in the background after response is sent.
    """
    # distributor can only confirm own or their sub_account orders
    user_role = UserRole(current_user.role)
    if user_role == UserRole.DISTRIBUTOR:
        order = await _get_order_or_404(db, order_id)
        if not order_service.can_view_order(current_user, order):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="총판은 본인 또는 하부 계정의 주문만 접수 확인할 수 있습니다.",
            )
    return await _confirm_and_start_pipeline(order_id, background_tasks, db, current_user)


@router.post("/{order_id}/reject", response_model=OrderResponse)
async def reject_order(
    order_id: int,
    body: OrderRejectRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Reject an order: submitted -> rejected. company_admin or system_admin only."""
    order = await _get_order_or_404(db, order_id)
    _check_company_scope(current_user, order)

    try:
        await order_service.reject_order(db, order, body.reason)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return await order_service.get_order_by_id(db, order_id)


@router.post("/{order_id}/hold", response_model=OrderResponse)
async def hold_order(
    order_id: int,
    body: OrderHoldRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Hold an order: submitted/payment_hold -> payment_hold. company_admin or system_admin only."""
    order = await _get_order_or_404(db, order_id)
    _check_company_scope(current_user, order)

    try:
        await order_service.hold_order(db, order, body.reason, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return await order_service.get_order_by_id(db, order_id)


@router.post("/{order_id}/release-hold", response_model=OrderResponse)
async def release_hold(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Release hold: payment_hold -> submitted. company_admin or system_admin only."""
    order = await _get_order_or_404(db, order_id)
    _check_company_scope(current_user, order)

    try:
        await order_service.release_hold(db, order, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return await order_service.get_order_by_id(db, order_id)


@router.post("/{order_id}/cancel", response_model=OrderResponse)
async def cancel_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Cancel an order: draft/submitted -> cancelled. company_admin or system_admin only."""
    order = await _get_order_or_404(db, order_id)
    _check_company_scope(current_user, order)

    try:
        await order_service.cancel_order(db, order, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return await order_service.get_order_by_id(db, order_id)


@router.post("/{order_id}/approve", response_model=OrderResponse)
async def approve_order(
    order_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN])
    ),
):
    """Approve order: submitted -> payment_confirmed. Alias for confirm-payment."""
    return await _confirm_and_start_pipeline(order_id, background_tasks, db, current_user)


@router.get("/{order_id}/items/export")
async def export_order_items(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export order items as an Excel file."""
    from openpyxl import Workbook

    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found",
        )
    if not order_service.can_view_order(current_user, order):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this order",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Items"

    is_sub_account = current_user.role == UserRole.SUB_ACCOUNT.value
    if is_sub_account:
        headers = ["Row #", "Product ID", "Quantity", "Status", "Result"]
    else:
        headers = [
            "Row #", "Product ID", "Quantity", "Unit Price",
            "Subtotal", "Status", "Result",
        ]
    ws.append(headers)

    for item in order.items:
        if is_sub_account:
            ws.append([
                item.row_number,
                item.product_id,
                item.quantity,
                item.status,
                item.result_message or "",
            ])
        else:
            ws.append([
                item.row_number,
                item.product_id,
                item.quantity,
                int(item.unit_price) if item.unit_price else 0,
                int(item.subtotal) if item.subtotal else 0,
                item.status,
                item.result_message or "",
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"order_{order.order_number}_items.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.patch("/{order_id}/deadline", response_model=OrderResponse)
async def update_deadline(
    order_id: int,
    body: DeadlineUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        RoleChecker([UserRole.SYSTEM_ADMIN, UserRole.COMPANY_ADMIN, UserRole.ORDER_HANDLER])
    ),
):
    """Update the deadline for an order."""
    order = await order_service.get_order_by_id(db, order_id)
    if order is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found",
        )

    try:
        updated = await order_service.update_deadline(db, order, body.deadline)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    return updated


# ---------------------------------------------------------------------------
# Background pipeline helpers
# ---------------------------------------------------------------------------

async def _dispatch_extraction_jobs_background(order_id: int) -> None:
    """Dispatch extraction jobs to keyword-worker in background.

    Called AFTER the main request committed pipeline_state records.
    Waits briefly to ensure the commit is visible, then dispatches.
    """
    import asyncio
    from app.core.database import async_session_factory

    # Wait for the main request's DB transaction to be committed.
    await asyncio.sleep(0.5)

    try:
        async with async_session_factory() as db:
            try:
                await pipeline_orchestrator.dispatch_pending_extraction_jobs(db)
                logger.info("Background extraction dispatch completed for order %s", order_id)
            except Exception:
                logger.exception("Background extraction dispatch failed for order %s", order_id)
    except Exception:
        logger.exception("Background dispatch session error for order %s", order_id)


async def _run_pipeline_for_order(order_id: int) -> None:
    """Run full pipeline start in background (legacy/fallback).

    Two-phase approach to avoid race conditions:
    Phase 1: Create DB records (pipeline_state, extraction_job) and COMMIT
    Phase 2: Dispatch to keyword-worker (worker can now see committed records)
    """
    import asyncio
    from app.core.database import async_session_factory

    # Wait briefly to ensure the main request's DB transaction has committed.
    await asyncio.sleep(0.5)

    # Phase 1: Create pipeline records and commit
    try:
        async with async_session_factory() as db:
            try:
                order = await order_service.get_order_by_id(db, order_id)
                if order is None:
                    logger.error("Background pipeline: order %s not found", order_id)
                    return
                await pipeline_orchestrator.start_pipeline_for_order(db, order)
                await db.commit()
                logger.info("Background pipeline records committed for order %s", order_id)
            except Exception:
                await db.rollback()
                logger.exception("Background pipeline failed for order %s", order_id)
                return
    except Exception:
        logger.exception("Background pipeline session error for order %s", order_id)
        return

    # Phase 2: Dispatch extraction jobs (AFTER commit, so worker can see records)
    try:
        async with async_session_factory() as db:
            try:
                await pipeline_orchestrator.dispatch_pending_extraction_jobs(db)
            except Exception:
                logger.exception("Background pipeline dispatch failed for order %s", order_id)
    except Exception:
        logger.exception("Background pipeline dispatch session error for order %s", order_id)
