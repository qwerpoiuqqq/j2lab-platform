"""Campaign upload service: Excel parsing, validation, campaign creation."""

from __future__ import annotations

import io
import logging
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.campaign import Campaign
from app.schemas.campaign import CampaignCreate
from app.services import campaign_service
from app.services.worker_clients import dispatch_campaign_registration, WorkerDispatchError

logger = logging.getLogger(__name__)

EXPECTED_HEADERS = ["대행사", "계정ID", "플레이스URL", "상호명", "캠페인타입", "시작일", "종료일", "일일한도", "키워드"]

CAMPAIGN_TYPE_MAP = {
    # 기본 영문 키
    "traffic": "traffic",
    "save": "save",
    "landmark": "landmark",
    # 한글 → 영문 매핑 (QCA 6개 타입)
    "트래픽": "트래픽",
    "트래픽1": "트래픽1",
    "저장하기": "저장하기",
    "저장하기1": "저장하기1",
    "명소": "명소",
    "공유+길찾기+트래픽": "공유+길찾기+트래픽",
    # 레거시 매핑
    "랜드마크": "landmark",
}


def parse_excel(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse Excel file and return headers + rows as dicts."""
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    if len(raw_rows) < 2:
        return [], []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(raw_rows[0])]
    rows = []
    for row in raw_rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val
        rows.append(row_dict)

    return headers, rows


def validate_row(row: dict, row_number: int) -> tuple[dict, list[str]]:
    """Validate a single row. Returns (normalized_data, errors)."""
    errors = []

    place_url = str(row.get("플레이스URL", "") or "").strip()
    if not place_url:
        errors.append("플레이스URL은 필수입니다.")
    elif not (place_url.startswith("http://") or place_url.startswith("https://")):
        errors.append("플레이스URL이 유효하지 않습니다.")

    campaign_type_raw = str(row.get("캠페인타입", "") or "").strip()
    campaign_type = CAMPAIGN_TYPE_MAP.get(campaign_type_raw, "")
    if not campaign_type:
        errors.append(f"캠페인타입 '{campaign_type_raw}'이(가) 유효하지 않습니다.")

    start_date_raw = row.get("시작일", "")
    end_date_raw = row.get("종료일", "")
    start_date = _parse_date(start_date_raw)
    end_date = _parse_date(end_date_raw)
    if not start_date:
        errors.append("시작일 형식이 잘못되었습니다.")
    if not end_date:
        errors.append("종료일 형식이 잘못되었습니다.")
    if start_date and end_date and end_date < start_date:
        errors.append("종료일이 시작일보다 앞입니다.")

    daily_limit_raw = row.get("일일한도", "")
    try:
        daily_limit = int(daily_limit_raw) if daily_limit_raw else 0
        if daily_limit < 1:
            errors.append("일일한도는 1 이상이어야 합니다.")
    except (ValueError, TypeError):
        daily_limit = 0
        errors.append("일일한도가 숫자가 아닙니다.")

    keywords_raw = str(row.get("키워드", "") or "").strip()
    keywords = [k.strip() for k in keywords_raw.split(",") if k.strip()] if keywords_raw else []
    if not keywords:
        errors.append("키워드가 비어있습니다.")

    normalized = {
        "agency_name": str(row.get("대행사", "") or "").strip(),
        "account_user_id": str(row.get("계정ID", "") or "").strip(),
        "place_url": place_url,
        "place_name": str(row.get("상호명", "") or "").strip(),
        "campaign_type": campaign_type,
        "start_date": start_date.isoformat() if start_date else "",
        "end_date": end_date.isoformat() if end_date else "",
        "daily_limit": daily_limit,
        "keywords": keywords,
    }

    return normalized, errors


async def detect_extension(
    db: AsyncSession, place_url: str, campaign_type: str
) -> tuple[bool, int | None, str | None]:
    """Check if this is an extension of an existing campaign."""
    result = await db.execute(
        select(Campaign).where(
            Campaign.place_url == place_url,
            Campaign.campaign_type == campaign_type,
            Campaign.status.in_(["active", "completed"]),
        ).order_by(Campaign.created_at.desc()).limit(1)
    )
    existing = result.scalar_one_or_none()
    if existing:
        return True, existing.id, existing.campaign_code
    return False, None, None


async def create_campaigns_from_upload(
    db: AsyncSession,
    items: list[dict],
) -> list[dict]:
    """Create campaigns from validated upload data."""
    results = []
    for item in items:
        try:
            campaign = await campaign_service.create_campaign(
                db,
                CampaignCreate(
                    place_url=item["place_url"],
                    place_name=item.get("place_name", ""),
                    campaign_type=item["campaign_type"],
                    start_date=item["start_date"],
                    end_date=item["end_date"],
                    daily_limit=item["daily_limit"],
                    agency_name=item.get("agency_name"),
                ),
            )
            # Add keywords
            if item.get("keywords"):
                await campaign_service.add_keywords_to_pool(
                    db, campaign_id=campaign.id, keywords=item["keywords"]
                )

            # Best-effort dispatch
            try:
                await dispatch_campaign_registration(campaign_id=campaign.id)
            except WorkerDispatchError:
                logger.warning("Failed to dispatch campaign %s registration", campaign.id)

            results.append({"campaign_id": campaign.id, "status": "created"})
        except Exception as e:
            logger.error("Failed to create campaign from upload: %s", e)
            results.append({"error": str(e)})

    return results


def _parse_date(val) -> date | None:
    """Parse various date formats."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if val is None:
        return None
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        pass
    try:
        return datetime.strptime(str(val).strip(), "%Y/%m/%d").date()
    except (ValueError, TypeError):
        return None
